from datetime import date
from decimal import Decimal
from io import BytesIO

import openpyxl
from django.test import TestCase
from django.urls import reverse

from accounts.models import User
from academics.models import AcademicYear, Class, Term
from core.models import FeeCategory, Gender, GradeLevel, Organization, TermChoices, UserRole
from finance.models import Invoice, InvoiceItem
from students.models import Student


class InvoiceReportRegressionTests(TestCase):
    def setUp(self):
        self.organization = Organization.objects.create(name="Test School", code="TEST")
        self.user = User.objects.create_user(
            email="accountant@example.com",
            password="password123",
            first_name="Report",
            last_name="User",
            role=UserRole.ACCOUNTANT,
            organization=self.organization,
        )
        self.client.force_login(self.user)

        self.academic_year = AcademicYear.objects.create(
            organization=self.organization,
            year=2026,
            start_date=date(2026, 1, 1),
            end_date=date(2026, 12, 31),
            is_current=True,
        )
        self.term = Term.objects.create(
            organization=self.organization,
            academic_year=self.academic_year,
            term=TermChoices.TERM_1,
            start_date=date(2026, 1, 1),
            end_date=date(2026, 4, 30),
            is_current=True,
        )
        self.student_class = Class.objects.create(
            organization=self.organization,
            name="Grade 4 North",
            grade_level=GradeLevel.GRADE_4,
            academic_year=self.academic_year,
        )
        self.student = Student.objects.create(
            organization=self.organization,
            admission_number="PWA1001",
            admission_date=date(2026, 1, 5),
            first_name="Ada",
            middle_name="Grace",
            last_name="Otieno",
            gender=Gender.FEMALE,
            date_of_birth=date(2016, 6, 10),
            current_class=self.student_class,
            status="active",
        )
        self.invoice = Invoice.objects.create(
            organization=self.organization,
            student=self.student,
            term=self.term,
            subtotal=Decimal("2000.00"),
            total_amount=Decimal("2000.00"),
            balance_bf=Decimal("150.00"),
            prepayment=Decimal("75.00"),
            issue_date=date(2026, 1, 10),
            due_date=date(2026, 1, 31),
            generated_by=self.user,
        )
        self._create_item(FeeCategory.TUITION, "Tuition", "1000.00")
        self._create_item(FeeCategory.EXAMINATION, "Exam Fee", "500.00")
        self._create_item(FeeCategory.ADMISSION, "Admission Fee", "300.00")
        self._create_item(FeeCategory.OTHER, "Lab Manual", "200.00")

    def _create_item(self, category, description, amount):
        return InvoiceItem.objects.create(
            invoice=self.invoice,
            description=description,
            category=category,
            amount=Decimal(amount),
            net_amount=Decimal(amount),
        )

    def _summary_params(self):
        return {
            "academic_year": str(self.academic_year.pk),
            "term": self.term.term,
            "show_zero_rows": "on",
        }

    def _detailed_params(self, **extra):
        params = {
            "academic_year": str(self.academic_year.pk),
            "term": self.term.term,
        }
        params.update(extra)
        return params

    def test_invoice_detailed_report_uses_canonical_category_choices_and_filters(self):
        response = self.client.get(reverse("reports:invoice_detailed_report"), self._detailed_params())

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Examination")
        self.assertContains(response, "Admission")
        self.assertContains(response, "Other: Lab Manual")

        filtered = self.client.get(
            reverse("reports:invoice_detailed_report"),
            self._detailed_params(category=[FeeCategory.EXAMINATION, FeeCategory.ADMISSION]),
        )

        self.assertContains(filtered, "Examination")
        self.assertContains(filtered, "Admission")
        self.assertNotContains(filtered, "Lab Manual")

    def test_invoice_detailed_export_renders_other_description_and_canonical_labels(self):
        response = self.client.get(
            reverse("reports:invoice_detailed_report_export_excel"),
            self._detailed_params(
                category=[FeeCategory.EXAMINATION, FeeCategory.ADMISSION, f"{FeeCategory.OTHER}:Lab Manual"]
            ),
        )

        self.assertEqual(response.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        sheet = workbook.active
        exported_categories = [sheet.cell(row=row, column=6).value for row in range(6, 9)]

        self.assertEqual(exported_categories, ["Examination", "Admission", "Lab Manual"])

    def test_invoice_summary_reports_show_current_term_prepayments_as_positive(self):
        response = self.client.get(reverse("reports:invoice_summary_report"), self._summary_params())

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Current Term Adjustments")
        self.assertContains(response, "Prepayments (KES)")
        self.assertContains(response, "KES 75.00")
        self.assertContains(response, "KES 150.00")
        self.assertContains(response, "Examination")
        self.assertContains(response, "Admission")
        self.assertContains(response, "Educational Activities")

        export_response = self.client.get(
            reverse("reports:invoice_summary_report_export_excel"),
            self._summary_params(),
        )
        self.assertEqual(export_response.status_code, 200)

        workbook = openpyxl.load_workbook(BytesIO(export_response.content))
        sheet = workbook.active
        category_labels = [sheet.cell(row=row, column=1).value for row in range(5, 15)]

        self.assertEqual(sheet.cell(row=5, column=6).value, "Prepayments (KES)")
        self.assertEqual(sheet.cell(row=10, column=1).value, "Current Term Adjustments")
        self.assertEqual(sheet.cell(row=10, column=5).value, 150.0)
        self.assertEqual(sheet.cell(row=10, column=6).value, 75.0)
        self.assertEqual(sheet.cell(row=11, column=6).value, 75.0)
        self.assertIn("Educational Activities", category_labels)
