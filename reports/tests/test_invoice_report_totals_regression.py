from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

import openpyxl
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from django.test import RequestFactory, TestCase
from django.urls import reverse
from django.utils import timezone

from academics.models import AcademicYear, Class, Term
from core.models import GradeLevel, Organization, PaymentMethod, PaymentSource, PaymentStatus, StreamChoices, TermChoices
from finance.models import Invoice, InvoiceItem
from payments.models import Payment, PaymentAllocation
from reports.views import InvoiceDetailedReportView, InvoiceReportView
from reports.views_exports import InvoiceSummaryReportExcelView
from students.models import Student


class InvoiceReportTotalsRegressionTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.organization = Organization.objects.create(name='PCEA Wendani Academy', code='PWA')
        self.user = user_model.objects.create_user(
            email='invoice-regression@example.com',
            password='testpass123',
            first_name='Invoice',
            last_name='Reporter',
            organization=self.organization,
        )
        self.factory = RequestFactory()

        self.academic_year = AcademicYear.objects.create(
            organization=self.organization,
            year=2026,
            start_date=date(2026, 1, 1),
            end_date=date(2026, 12, 31),
            is_current=True,
        )
        self.term = Term.objects.create(
            organization=self.organization,
            academic_year=self.academic_year,
            term=TermChoices.TERM_1,
            start_date=date(2026, 1, 1),
            end_date=date(2026, 4, 30),
            is_current=True,
        )
        self.student_class = Class.objects.create(
            organization=self.organization,
            name='Grade 5 West',
            grade_level=GradeLevel.GRADE_5,
            stream=StreamChoices.WEST,
            academic_year=self.academic_year,
        )
        self.student = Student.objects.create(
            organization=self.organization,
            admission_number='PWA5001',
            admission_date=date(2026, 1, 10),
            first_name='Nia',
            last_name='Totals',
            gender='F',
            date_of_birth=date(2015, 3, 1),
            current_class=self.student_class,
            status='active',
        )

        invoice = Invoice.objects.create(
            organization=self.organization,
            invoice_number='INV-2026-1001',
            student=self.student,
            term=self.term,
            subtotal=Decimal('1800.00'),
            total_amount=Decimal('1800.00'),
            amount_paid=Decimal('0.00'),
            balance=Decimal('1800.00'),
            status='partially_paid',
            issue_date=date(2026, 1, 15),
            due_date=date(2026, 2, 15),
        )
        tuition_item = InvoiceItem.objects.create(
            invoice=invoice,
            description='Tuition',
            category='tuition',
            amount=Decimal('1200.00'),
            net_amount=Decimal('1200.00'),
        )
        meals_item = InvoiceItem.objects.create(
            invoice=invoice,
            description='Meals',
            category='meals',
            amount=Decimal('600.00'),
            net_amount=Decimal('600.00'),
        )

        payment = Payment.objects.create(
            payment_reference='PAY-INV-001',
            organization=self.organization,
            student=self.student,
            invoice=invoice,
            amount=Decimal('1000.00'),
            payment_method=PaymentMethod.MOBILE_MONEY,
            payment_source=PaymentSource.MPESA,
            status=PaymentStatus.COMPLETED,
            payment_date=timezone.make_aware(datetime(2026, 1, 20, 10, 0), timezone.get_current_timezone()),
        )
        PaymentAllocation.objects.create(payment=payment, invoice_item=tuition_item, amount=Decimal('700.00'))
        PaymentAllocation.objects.create(payment=payment, invoice_item=meals_item, amount=Decimal('300.00'))

        self.summary_url = reverse('reports:invoice_summary_report')
        self.summary_excel_url = reverse('reports:invoice_summary_report_export_excel')
        self.detailed_url = reverse('reports:invoice_detailed_report')

    def _params(self):
        return {
            'academic_year': self.academic_year.pk,
            'term': self.term.term,
        }

    def _get_summary_html_totals(self, params):
        request = self.factory.get(self.summary_url, data=params)
        request.user = self.user
        request.organization = self.organization

        with patch('reports.views.render') as mock_render:
            mock_render.side_effect = lambda request, template_name, context: HttpResponse('rendered')
            response = InvoiceReportView.as_view()(request)

        self.assertEqual(response.status_code, 200)
        return mock_render.call_args.args[2]['totals']

    def _get_summary_excel_totals(self, params):
        request = self.factory.get(self.summary_excel_url, data=params)
        request.user = self.user
        request.organization = self.organization

        response = InvoiceSummaryReportExcelView.as_view()(request)
        self.assertEqual(response.status_code, 200)

        workbook = openpyxl.load_workbook(BytesIO(response.content))
        sheet = workbook.active

        totals_row = None
        row_num = 6
        while row_num <= sheet.max_row:
            if sheet.cell(row=row_num, column=1).value == 'TOTALS':
                totals_row = row_num
                break
            row_num += 1

        self.assertIsNotNone(totals_row, 'Expected TOTALS row in invoice summary export workbook')
        return {
            'billed': Decimal(str(sheet.cell(row=totals_row, column=2).value or 0)),
            'collected': Decimal(str(sheet.cell(row=totals_row, column=3).value or 0)),
            'outstanding': Decimal(str(sheet.cell(row=totals_row, column=4).value or 0)),
        }

    def _get_detailed_html_totals(self, params):
        request = self.factory.get(self.detailed_url, data=params)
        request.user = self.user
        request.organization = self.organization

        with patch('reports.views.render') as mock_render:
            mock_render.side_effect = lambda request, template_name, context: HttpResponse('rendered')
            response = InvoiceDetailedReportView.as_view()(request)

        self.assertEqual(response.status_code, 200)
        totals = mock_render.call_args.args[2]['totals']
        return {
            'billed': totals['total_billed'],
            'collected': totals['total_paid'],
            'outstanding': totals['total_balance'],
        }

    def test_summary_html_excel_and_detailed_totals_match_for_same_filters(self):
        params = self._params()

        summary_html_totals = self._get_summary_html_totals(params)
        summary_excel_totals = self._get_summary_excel_totals(params)
        detailed_totals = self._get_detailed_html_totals(params)

        normalized_summary_html = {
            'billed': summary_html_totals['billed'],
            'collected': summary_html_totals['collected'],
            'outstanding': summary_html_totals['outstanding'],
        }

        self.assertEqual(normalized_summary_html, summary_excel_totals)
        self.assertEqual(normalized_summary_html, detailed_totals)
