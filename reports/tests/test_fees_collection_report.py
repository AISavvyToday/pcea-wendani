from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

import openpyxl
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from django.test import RequestFactory, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from academics.models import AcademicYear, Class, Term
from core.models import (
    GradeLevel,
    Organization,
    PaymentMethod,
    PaymentSource,
    PaymentStatus,
    StreamChoices,
    TermChoices,
)
from finance.models import Invoice, InvoiceItem
from payments.models import Payment, PaymentAllocation
from reports.views import FeesCollectionReportView
from reports.views_exports import FeesCollectionExcelView, FeesCollectionPDFView
from students.models import Student


@override_settings(STATICFILES_STORAGE='django.contrib.staticfiles.storage.StaticFilesStorage')
class FeesCollectionReportRegressionTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.organization = Organization.objects.create(
            name='PCEA Wendani Academy',
            code='PWA',
        )
        self.user = user_model.objects.create_user(
            email='reporter@example.com',
            password='testpass123',
            first_name='Report',
            last_name='User',
            organization=self.organization,
        )
        self.client.force_login(self.user)
        self.factory = RequestFactory()

        self.academic_year = AcademicYear.objects.create(
            organization=self.organization,
            year=2026,
            start_date=date(2026, 1, 1),
            end_date=date(2026, 12, 31),
            is_current=True,
        )
        self.term = Term.objects.create(
            organization=self.organization,
            academic_year=self.academic_year,
            term=TermChoices.TERM_1,
            start_date=date(2026, 1, 1),
            end_date=date(2026, 4, 30),
            is_current=True,
        )
        self.student_class = Class.objects.create(
            organization=self.organization,
            name='Grade 4 East',
            grade_level=GradeLevel.GRADE_4,
            stream=StreamChoices.EAST,
            academic_year=self.academic_year,
        )
        self.student = Student.objects.create(
            organization=self.organization,
            admission_number='PWA4001',
            admission_date=date(2026, 1, 10),
            first_name='Amani',
            middle_name='Test',
            last_name='Student',
            gender='F',
            date_of_birth=date(2016, 5, 1),
            current_class=self.student_class,
            status='active',
        )
        self.invoice = Invoice.objects.create(
            organization=self.organization,
            invoice_number='INV-2026-0001',
            student=self.student,
            term=self.term,
            subtotal=Decimal('1000.00'),
            total_amount=Decimal('1000.00'),
            amount_paid=Decimal('0.00'),
            balance=Decimal('1000.00'),
            status='partially_paid',
            issue_date=date(2026, 1, 10),
            due_date=date(2026, 1, 31),
        )
        self.tuition_item = InvoiceItem.objects.create(
            invoice=self.invoice,
            description='Tuition',
            category='tuition',
            amount=Decimal('500.00'),
            net_amount=Decimal('500.00'),
        )
        self.chess_item = InvoiceItem.objects.create(
            invoice=self.invoice,
            description='Chess Club',
            category='other',
            amount=Decimal('550.00'),
            net_amount=Decimal('550.00'),
        )
        self.scout_item = InvoiceItem.objects.create(
            invoice=self.invoice,
            description='Scout Trip',
            category='other',
            amount=Decimal('150.00'),
            net_amount=Decimal('150.00'),
        )

        self.manual_payment = self.create_payment(
            payment_reference='PAY-HTML-001',
            receipt_number='RCP-HTML-001',
            transaction_reference='TXN-MANUAL-001',
            amount=Decimal('500.00'),
            payment_source=PaymentSource.MPESA,
            payment_method=PaymentMethod.MOBILE_MONEY,
            payment_date=self.make_aware(2026, 1, 15, 0, 1),
            allocations=[
                (self.chess_item, Decimal('300.00')),
                (self.tuition_item, Decimal('200.00')),
            ],
        )
        self.bank_payment = self.create_payment(
            payment_reference='PAY-HTML-002',
            receipt_number='',
            transaction_reference='TXN-BANK-002',
            amount=Decimal('250.00'),
            payment_source=PaymentSource.COOP_BANK,
            payment_method=PaymentMethod.BANK_DEPOSIT,
            payment_date=self.make_aware(2026, 1, 15, 23, 59),
            allocations=[
                (self.chess_item, Decimal('250.00')),
            ],
        )
        self.transaction_only_payment = self.create_payment(
            payment_reference='PAY-HTML-003',
            receipt_number='',
            transaction_reference='TXN-ONLY-003',
            amount=Decimal('150.00'),
            payment_source=PaymentSource.EQUITY_BANK,
            payment_method=PaymentMethod.BANK_DEPOSIT,
            payment_date=self.make_aware(2026, 1, 16, 12, 0),
            allocations=[
                (self.scout_item, Decimal('150.00')),
            ],
        )
        Payment.objects.filter(pk=self.transaction_only_payment.pk).update(payment_reference='', receipt_number='')
        self.transaction_only_payment.refresh_from_db()

        self.html_url = reverse('reports:fees_collection_report')
        self.xlsx_url = reverse('reports:fees_collection_export_excel')
        self.pdf_url = reverse('reports:fees_collection_export_pdf')

    def make_aware(self, year, month, day, hour, minute):
        return timezone.make_aware(datetime(year, month, day, hour, minute), timezone.get_current_timezone())

    def create_payment(
        self,
        *,
        payment_reference,
        receipt_number,
        transaction_reference,
        amount,
        payment_source,
        payment_method,
        payment_date,
        allocations,
    ):
        payment = Payment.objects.create(
            payment_reference=payment_reference,
            organization=self.organization,
            student=self.student,
            invoice=self.invoice,
            amount=amount,
            payment_method=payment_method,
            payment_source=payment_source,
            status=PaymentStatus.COMPLETED,
            payment_date=payment_date,
            transaction_reference=transaction_reference,
            receipt_number=receipt_number,
        )
        if receipt_number == '':
            Payment.objects.filter(pk=payment.pk).update(receipt_number='')
            payment.refresh_from_db()

        for invoice_item, allocation_amount in allocations:
            PaymentAllocation.objects.create(
                payment=payment,
                invoice_item=invoice_item,
                amount=allocation_amount,
            )
        return payment

    def get_pdf_html(self, params):
        with patch('reports.views_exports.HTML') as mock_html:
            mock_html.return_value.write_pdf.return_value = b'%PDF-1.4 test'
            response = self.make_view_request(FeesCollectionPDFView, self.pdf_url, params)
            self.assertEqual(response.status_code, 200)
            return mock_html.call_args.kwargs['string']

    def get_excel_rows(self, params):
        response = self.make_view_request(FeesCollectionExcelView, self.xlsx_url, params)
        self.assertEqual(response.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        sheet = workbook.active
        rows = []
        row_num = 6
        while sheet.cell(row=row_num, column=1).value:
            rows.append(
                {
                    'date': sheet.cell(row=row_num, column=1).value,
                    'reference': sheet.cell(row=row_num, column=2).value,
                    'student': sheet.cell(row=row_num, column=3).value,
                    'admission': sheet.cell(row=row_num, column=4).value,
                    'class': sheet.cell(row=row_num, column=5).value,
                    'amount': Decimal(str(sheet.cell(row=row_num, column=6).value)),
                }
            )
            row_num += 1
        return rows

    def make_view_request(self, view_class, url, params=None):
        request = self.factory.get(url, data=params or {})
        request.user = self.user
        request.organization = self.organization
        return view_class.as_view()(request)

    def get_html_context(self, params=None):
        request = self.factory.get(self.html_url, data=params or {})
        request.user = self.user
        request.organization = self.organization

        with patch('reports.views.render') as mock_render:
            mock_render.side_effect = lambda request, template_name, context: HttpResponse('rendered')
            response = FeesCollectionReportView.as_view()(request)

        self.assertEqual(response.status_code, 200)
        return mock_render.call_args.args[2]

    def test_dynamic_other_item_choices_are_available_on_form(self):
        context = self.get_html_context()

        category_choices = context['form'].fields['category'].choices
        self.assertIn(('other:Chess Club', 'Other: Chess Club'), category_choices)
        self.assertIn(('other:Scout Trip', 'Other: Scout Trip'), category_choices)
        self.assertIn(('other', 'Other'), category_choices)
        self.assertIn('Receipt / Ref', open('templates/reports/fees_collection_report.html').read())

    def test_html_report_filters_same_day_records_and_other_item_allocations(self):
        context = self.get_html_context(
            {
                'start_date': '2026-01-15',
                'end_date': '2026-01-15',
                'category': ['other:Chess Club'],
            },
        )

        rows = context['rows']
        self.assertEqual(len(rows), 2)
        self.assertEqual([row['reference'] for row in rows], ['RCP-HTML-001', 'PAY-HTML-002'])
        self.assertEqual([row['amount'] for row in rows], [Decimal('300.00'), Decimal('250.00')])
        self.assertEqual(context['summary']['total_collected'], Decimal('550.00'))
        self.assertEqual(context['filters']['category_labels'], ['Other: Chess Club'])
        self.assertNotIn('TXN-ONLY-003', [row['reference'] for row in rows])

    def test_html_pdf_and_excel_share_reference_and_amount_semantics(self):
        params = {
            'start_date': '2026-01-15',
            'end_date': '2026-01-16',
            'category': ['other:Chess Club', 'other:Scout Trip'],
        }

        html_context = self.get_html_context(params)
        html_rows = html_context['rows']
        self.assertEqual(
            [(row['reference'], row['amount']) for row in html_rows],
            [
                ('RCP-HTML-001', Decimal('300.00')),
                ('PAY-HTML-002', Decimal('250.00')),
                ('TXN-ONLY-003', Decimal('150.00')),
            ],
        )

        pdf_html = self.get_pdf_html(params)
        self.assertIn('Receipt / Ref', pdf_html)
        self.assertIn('RCP-HTML-001', pdf_html)
        self.assertIn('PAY-HTML-002', pdf_html)
        self.assertIn('TXN-ONLY-003', pdf_html)
        self.assertIn('Other: Chess Club, Other: Scout Trip', pdf_html)

        excel_rows = self.get_excel_rows(params)
        self.assertEqual(
            [(row['reference'], row['amount']) for row in excel_rows],
            [
                ('RCP-HTML-001', Decimal('300')),
                ('PAY-HTML-002', Decimal('250')),
                ('TXN-ONLY-003', Decimal('150')),
            ],
        )
