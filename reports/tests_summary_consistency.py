from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO

import openpyxl
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from accounts.models import User
from academics.models import AcademicYear, Class, Term
from core.models import (
    FeeCategory,
    Gender,
    GradeLevel,
    Organization,
    PaymentMethod,
    PaymentSource,
    PaymentStatus,
    TermChoices,
    UserRole,
)
from finance.models import Invoice, InvoiceItem
from payments.models import Payment, PaymentAllocation
from reports.report_utils import build_invoice_detailed_report_data, build_invoice_summary_report_data
from students.models import Student


@override_settings(STATICFILES_STORAGE='django.contrib.staticfiles.storage.StaticFilesStorage')
class InvoiceSummaryConsistencyTests(TestCase):
    def setUp(self):
        self.organization = Organization.objects.create(name="Consistency School", code="CONS")
        self.user = User.objects.create_user(
            email="consistency@example.com",
            password="password123",
            first_name="Consistency",
            last_name="Tester",
            role=UserRole.ACCOUNTANT,
            organization=self.organization,
        )
        self.client.force_login(self.user)

        self.academic_year = AcademicYear.objects.create(
            organization=self.organization,
            year=2026,
            start_date=date(2026, 1, 1),
            end_date=date(2026, 12, 31),
            is_current=True,
        )
        self.term = Term.objects.create(
            organization=self.organization,
            academic_year=self.academic_year,
            term=TermChoices.TERM_1,
            start_date=date(2026, 1, 1),
            end_date=date(2026, 4, 30),
            is_current=True,
        )
        school_class = Class.objects.create(
            organization=self.organization,
            name="Grade 3 Blue",
            grade_level=GradeLevel.GRADE_3,
            academic_year=self.academic_year,
        )
        self.student = Student.objects.create(
            organization=self.organization,
            admission_number="PWA2001",
            admission_date=date(2026, 1, 5),
            first_name="Lena",
            middle_name="A",
            last_name="Njeri",
            gender=Gender.FEMALE,
            date_of_birth=date(2017, 3, 14),
            current_class=school_class,
            status="active",
        )

        self.invoice = Invoice.objects.create(
            organization=self.organization,
            student=self.student,
            term=self.term,
            subtotal=Decimal("1500.00"),
            total_amount=Decimal("1500.00"),
            balance_bf=Decimal("200.00"),
            prepayment=Decimal("125.00"),
            issue_date=date(2026, 1, 12),
            due_date=date(2026, 1, 31),
            generated_by=self.user,
        )
        self.tuition_item = InvoiceItem.objects.create(
            invoice=self.invoice,
            description="Tuition",
            category=FeeCategory.TUITION,
            amount=Decimal("1000.00"),
            net_amount=Decimal("1000.00"),
        )
        self.other_item = InvoiceItem.objects.create(
            invoice=self.invoice,
            description="Workbook",
            category=FeeCategory.OTHER,
            amount=Decimal("500.00"),
            net_amount=Decimal("500.00"),
        )

        payment = Payment.objects.create(
            organization=self.organization,
            student=self.student,
            invoice=self.invoice,
            amount=Decimal("600.00"),
            payment_method=PaymentMethod.BANK_DEPOSIT,
            payment_source=PaymentSource.EQUITY_BANK,
            status=PaymentStatus.COMPLETED,
            payment_date=timezone.make_aware(datetime.combine(date(2026, 1, 20), time(9, 0))),
            received_by=self.user,
        )
        PaymentAllocation.objects.create(payment=payment, invoice_item=self.tuition_item, amount=Decimal("400.00"))
        PaymentAllocation.objects.create(payment=payment, invoice_item=self.other_item, amount=Decimal("200.00"))

    def _params(self):
        return {
            "academic_year": str(self.academic_year.pk),
            "term": self.term.term,
            "start_date": "2026-01-01",
            "end_date": "2026-01-31",
            "show_zero_rows": "on",
        }

    def test_summary_html_totals_match_excel_totals(self):
        html_response = self.client.get(reverse("reports:invoice_summary_report"), self._params())
        self.assertEqual(html_response.status_code, 200)
        html_totals = html_response.context["totals"]

        excel_response = self.client.get(reverse("reports:invoice_summary_report_export_excel"), self._params())
        self.assertEqual(excel_response.status_code, 200)
        sheet = openpyxl.load_workbook(BytesIO(excel_response.content)).active

        totals_row = None
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == "TOTALS":
                totals_row = row
                break
        self.assertIsNotNone(totals_row)

        self.assertEqual(Decimal(str(sheet.cell(row=totals_row, column=2).value)), html_totals["billed"])
        self.assertEqual(Decimal(str(sheet.cell(row=totals_row, column=3).value)), html_totals["collected"])
        self.assertEqual(Decimal(str(sheet.cell(row=totals_row, column=4).value)), html_totals["outstanding"])
        self.assertEqual(Decimal(str(sheet.cell(row=totals_row, column=5).value)), html_totals["balance_bf"])
        self.assertEqual(Decimal(str(sheet.cell(row=totals_row, column=6).value)), html_totals["prepayment_display"])

    def test_summary_totals_match_detailed_rollup_totals_for_same_filters(self):
        summary = build_invoice_summary_report_data(
            academic_year=self.academic_year,
            term=self.term.term,
            start_date=date(2026, 1, 1),
            end_date=date(2026, 1, 31),
            organization=self.organization,
            show_zero=True,
        )
        detailed = build_invoice_detailed_report_data(
            organization=self.organization,
            academic_year=self.academic_year,
            term=self.term.term,
            start_date=date(2026, 1, 1),
            end_date=date(2026, 1, 31),
            show_all=False,
        )

        self.assertEqual(summary["totals"]["billed"], detailed["totals"]["total_billed"])
        self.assertEqual(summary["totals"]["collected"], detailed["totals"]["total_paid"])
        self.assertEqual(summary["totals"]["outstanding"], detailed["totals"]["total_balance"])
        self.assertEqual(summary["totals"]["balance_bf"], Decimal("200.00"))
        self.assertEqual(summary["totals"]["prepayment"], Decimal("125.00"))
        self.assertEqual(summary["totals"]["prepayment_display"], Decimal("125.00"))
