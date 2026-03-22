from datetime import date
from decimal import Decimal
from io import BytesIO

import openpyxl
from django.test import TestCase
from django.urls import reverse

from accounts.models import User
from academics.models import AcademicYear, Class, Term
from core.models import GradeLevel, StreamChoices, TermChoices, UserRole
from core.models import Organization
from finance.models import Invoice
from students.models import Parent, Student, StudentParent


class OutstandingBalancesReportTests(TestCase):
    def setUp(self):
        self.organization = Organization.objects.create(
            name='Test School',
            code='TEST-SCHOOL',
        )
        self.user = User.objects.create_user(
            email='reports@example.com',
            password='password123',
            first_name='Report',
            last_name='User',
            role=UserRole.SCHOOL_ADMIN,
            is_staff=True,
from core.models import FeeCategory, Gender, GradeLevel, Organization, TermChoices, UserRole
from finance.models import Invoice, InvoiceItem
from students.models import Student


class InvoiceReportRegressionTests(TestCase):
    def setUp(self):
        self.organization = Organization.objects.create(name="Test School", code="TEST")
        self.user = User.objects.create_user(
            email="accountant@example.com",
            password="password123",
            first_name="Report",
            last_name="User",
            role=UserRole.ACCOUNTANT,
            organization=self.organization,
        )
        self.client.force_login(self.user)

        self.academic_year_2025 = AcademicYear.objects.create(
            organization=self.organization,
            year=2025,
            start_date=date(2025, 1, 1),
            end_date=date(2025, 12, 31),
        )
        self.academic_year_2026 = AcademicYear.objects.create(
        self.academic_year = AcademicYear.objects.create(
            organization=self.organization,
            year=2026,
            start_date=date(2026, 1, 1),
            end_date=date(2026, 12, 31),
        )
        self.term_2025 = Term.objects.create(
            organization=self.organization,
            academic_year=self.academic_year_2025,
            term=TermChoices.TERM_1,
            start_date=date(2025, 1, 1),
            end_date=date(2025, 4, 30),
        )
        self.term_2026 = Term.objects.create(
            organization=self.organization,
            academic_year=self.academic_year_2026,
            term=TermChoices.TERM_1,
            start_date=date(2026, 1, 1),
            end_date=date(2026, 4, 30),
        )
        self.student_class = Class.objects.create(
            organization=self.organization,
            name='Grade 4',
            grade_level=GradeLevel.GRADE_4,
            stream=StreamChoices.EAST,
            academic_year=self.academic_year_2025,
        )

    def create_student_with_invoice(self, *, admission_number, first_name, balance, issue_date, term):
        student = Student.objects.create(
            organization=self.organization,
            admission_number=admission_number,
            admission_date=issue_date,
            first_name=first_name,
            middle_name='',
            last_name='Student',
            gender='M',
            date_of_birth=date(2015, 1, 1),
            current_class=self.student_class,
            status='active',
        )
        Invoice.objects.create(
            organization=self.organization,
            invoice_number=f'INV-{admission_number}',
            student=student,
            term=term,
            subtotal=Decimal(balance),
            total_amount=Decimal(balance),
            amount_paid=Decimal('0.00'),
            balance=Decimal(balance),
            balance_bf=Decimal('0.00'),
            prepayment=Decimal('0.00'),
            issue_date=issue_date,
            due_date=issue_date,
        )
        return student

    def create_parent(self, *, first_name, last_name, phone_primary):
        return Parent.objects.create(
            organization=self.organization,
            first_name=first_name,
            last_name=last_name,
            phone_primary=phone_primary,
            relationship='guardian',
        )

    def get_html_rows_by_admission(self, **query_params):
        response = self.client.get(reverse('reports:outstanding_report'), query_params)
        self.assertEqual(response.status_code, 200)
        return {
            row['student__admission_number']: row
            for row in response.context['rows']
        }, response

    def get_excel_sheet(self, **query_params):
        response = self.client.get(reverse('reports:outstanding_report_export_excel'), query_params)
        self.assertEqual(response.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        return workbook.active

    def test_outstanding_report_prefers_primary_parent_contact(self):
        student = self.create_student_with_invoice(
            admission_number='ADM001',
            first_name='Primary',
            balance='1500.00',
            issue_date=date(2025, 2, 10),
            term=self.term_2025,
        )
        fallback_parent = self.create_parent(
            first_name='Fallback',
            last_name='Guardian',
            phone_primary='+254700000001',
        )
        primary_parent = self.create_parent(
            first_name='Primary',
            last_name='Guardian',
            phone_primary='+254700000002',
        )
        StudentParent.objects.create(
            student=student,
            parent=fallback_parent,
            relationship='guardian',
            is_primary=False,
        )
        StudentParent.objects.create(
            student=student,
            parent=primary_parent,
            relationship='guardian',
            is_primary=True,
        )

        rows, response = self.get_html_rows_by_admission(academic_year=self.academic_year_2025.pk)

        self.assertEqual(
            rows['ADM001']['parent_contact'],
            'Primary Guardian — +254700000002',
        )
        self.assertContains(response, 'Parent / Guardian Contact')

        sheet = self.get_excel_sheet(academic_year=self.academic_year_2025.pk)
        self.assertEqual(sheet['E5'].value, 'Parent / Guardian Contact')
        self.assertEqual(sheet['E6'].value, 'Primary Guardian — +254700000002')

    def test_outstanding_report_falls_back_to_first_linked_parent_when_no_primary(self):
        student = self.create_student_with_invoice(
            admission_number='ADM002',
            first_name='Fallback',
            balance='2400.00',
            issue_date=date(2025, 2, 11),
            term=self.term_2025,
        )
        first_parent = self.create_parent(
            first_name='First',
            last_name='Parent',
            phone_primary='+254700000003',
        )
        second_parent = self.create_parent(
            first_name='Second',
            last_name='Parent',
            phone_primary='+254700000004',
        )
        StudentParent.objects.create(
            student=student,
            parent=first_parent,
            relationship='guardian',
            is_primary=False,
        )
        StudentParent.objects.create(
            student=student,
            parent=second_parent,
            relationship='guardian',
            is_primary=False,
        )

        rows, _ = self.get_html_rows_by_admission(academic_year=self.academic_year_2025.pk)

        self.assertEqual(
            rows['ADM002']['parent_contact'],
            'First Parent — +254700000003',
        )

    def test_outstanding_report_uses_placeholder_when_student_has_no_linked_parents(self):
        self.create_student_with_invoice(
            admission_number='ADM003',
            first_name='NoParent',
            balance='3600.00',
            issue_date=date(2025, 2, 12),
            term=self.term_2025,
        )

        rows, _ = self.get_html_rows_by_admission(academic_year=self.academic_year_2025.pk)

        self.assertEqual(rows['ADM003']['parent_contact'], '—')

    def test_date_range_only_filter_matches_html_and_excel_exports_across_academic_years(self):
        in_range_2025 = self.create_student_with_invoice(
            admission_number='ADM004',
            first_name='RangeOne',
            balance='4700.00',
            issue_date=date(2025, 3, 15),
            term=self.term_2025,
        )
        in_range_2026 = self.create_student_with_invoice(
            admission_number='ADM005',
            first_name='RangeTwo',
            balance='5800.00',
            issue_date=date(2026, 3, 20),
            term=self.term_2026,
        )
        self.create_student_with_invoice(
            admission_number='ADM006',
            first_name='Outside',
            balance='6900.00',
            issue_date=date(2026, 5, 1),
            term=self.term_2026,
        )

        parent_one = self.create_parent(
            first_name='Year',
            last_name='One',
            phone_primary='+254700000005',
        )
        parent_two = self.create_parent(
            first_name='Year',
            last_name='Two',
            phone_primary='+254700000006',
        )
        StudentParent.objects.create(
            student=in_range_2025,
            parent=parent_one,
            relationship='guardian',
            is_primary=True,
        )
        StudentParent.objects.create(
            student=in_range_2026,
            parent=parent_two,
            relationship='guardian',
            is_primary=True,
        )

        params = {
            'start_date': '2025-03-01',
            'end_date': '2026-03-31',
        }
        rows, response = self.get_html_rows_by_admission(**params)

        self.assertEqual(set(rows.keys()), {'ADM004', 'ADM005'})
        self.assertContains(
            response,
            'Without an academic year, the date range applies across all invoices.',
        )

        sheet = self.get_excel_sheet(**params)
        excel_rows = {
            sheet.cell(row=row_idx, column=2).value: sheet.cell(row=row_idx, column=5).value
            for row_idx in range(6, sheet.max_row)
            if sheet.cell(row=row_idx, column=2).value != 'TOTALS'
        }

        self.assertEqual(set(excel_rows.keys()), {'ADM004', 'ADM005'})
        self.assertEqual(excel_rows['ADM004'], 'Year One — +254700000005')
        self.assertEqual(excel_rows['ADM005'], 'Year Two — +254700000006')
            is_current=True,
        )
        self.term = Term.objects.create(
            organization=self.organization,
            academic_year=self.academic_year,
            term=TermChoices.TERM_1,
            start_date=date(2026, 1, 1),
            end_date=date(2026, 4, 30),
            is_current=True,
        )
        self.student_class = Class.objects.create(
            organization=self.organization,
            name="Grade 4 North",
            grade_level=GradeLevel.GRADE_4,
            academic_year=self.academic_year,
        )
        self.student = Student.objects.create(
            organization=self.organization,
            admission_number="PWA1001",
            admission_date=date(2026, 1, 5),
            first_name="Ada",
            middle_name="Grace",
            last_name="Otieno",
            gender=Gender.FEMALE,
            date_of_birth=date(2016, 6, 10),
            current_class=self.student_class,
            status="active",
        )
        self.invoice = Invoice.objects.create(
            organization=self.organization,
            student=self.student,
            term=self.term,
            subtotal=Decimal("2000.00"),
            total_amount=Decimal("2000.00"),
            balance_bf=Decimal("150.00"),
            prepayment=Decimal("75.00"),
            issue_date=date(2026, 1, 10),
            due_date=date(2026, 1, 31),
            generated_by=self.user,
        )
        self._create_item(FeeCategory.TUITION, "Tuition", "1000.00")
        self._create_item(FeeCategory.EXAMINATION, "Exam Fee", "500.00")
        self._create_item(FeeCategory.ADMISSION, "Admission Fee", "300.00")
        self._create_item(FeeCategory.OTHER, "Lab Manual", "200.00")

    def _create_item(self, category, description, amount):
        return InvoiceItem.objects.create(
            invoice=self.invoice,
            description=description,
            category=category,
            amount=Decimal(amount),
            net_amount=Decimal(amount),
        )

    def _summary_params(self):
        return {
            "academic_year": str(self.academic_year.pk),
            "term": self.term.term,
            "show_zero_rows": "on",
        }

    def _detailed_params(self, **extra):
        params = {
            "academic_year": str(self.academic_year.pk),
            "term": self.term.term,
        }
        params.update(extra)
        return params

    def test_invoice_detailed_report_uses_canonical_category_choices_and_filters(self):
        response = self.client.get(reverse("reports:invoice_detailed_report"), self._detailed_params())

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Examination")
        self.assertContains(response, "Admission")
        self.assertContains(response, "Other: Lab Manual")

        filtered = self.client.get(
            reverse("reports:invoice_detailed_report"),
            self._detailed_params(category=[FeeCategory.EXAMINATION, FeeCategory.ADMISSION]),
        )

        self.assertContains(filtered, "Examination")
        self.assertContains(filtered, "Admission")
        self.assertNotContains(filtered, "Lab Manual")

    def test_invoice_detailed_export_renders_other_description_and_canonical_labels(self):
        response = self.client.get(
            reverse("reports:invoice_detailed_report_export_excel"),
            self._detailed_params(
                category=[FeeCategory.EXAMINATION, FeeCategory.ADMISSION, f"{FeeCategory.OTHER}:Lab Manual"]
            ),
        )

        self.assertEqual(response.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        sheet = workbook.active
        exported_categories = [sheet.cell(row=row, column=6).value for row in range(6, 9)]

        self.assertEqual(exported_categories, ["Examination", "Admission", "Lab Manual"])

    def test_invoice_summary_reports_show_current_term_prepayments_as_positive(self):
        response = self.client.get(reverse("reports:invoice_summary_report"), self._summary_params())

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Current Term Adjustments")
        self.assertContains(response, "Prepayments (KES)")
        self.assertContains(response, "KES 75.00")
        self.assertContains(response, "KES 150.00")
        self.assertContains(response, "Examination")
        self.assertContains(response, "Admission")

        export_response = self.client.get(
            reverse("reports:invoice_summary_report_export_excel"),
            self._summary_params(),
        )
        self.assertEqual(export_response.status_code, 200)

        workbook = openpyxl.load_workbook(BytesIO(export_response.content))
        sheet = workbook.active

        self.assertEqual(sheet.cell(row=5, column=6).value, "Prepayments (KES)")
        self.assertEqual(sheet.cell(row=10, column=1).value, "Current Term Adjustments")
        self.assertEqual(sheet.cell(row=10, column=5).value, 150.0)
        self.assertEqual(sheet.cell(row=10, column=6).value, 75.0)
        self.assertEqual(sheet.cell(row=11, column=6).value, 75.0)
