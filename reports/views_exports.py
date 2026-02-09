# reports/views_exports.py
import io
from decimal import Decimal
from datetime import datetime
from decimal import Decimal
from django.db.models import ExpressionWrapper, DecimalField
from django.shortcuts import render
from django.template.loader import render_to_string
from django.http import HttpResponse, HttpResponseBadRequest
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import View
from django.conf import settings
from django.db.models import Sum, Value, Q
from django.db.models.functions import Coalesce

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# WeasyPrint for PDF
from weasyprint import HTML

# Import models used by the reports
from finance.models import Invoice, InvoiceItem
from payments.models import Payment, PaymentAllocation
from academics.models import AcademicYear, Term
from transport.models import TransportRoute
from core.models import InvoiceStatus
from core.mixins import OrganizationFilterMixin
from students.models import Student


# ---------- Helpers ----------
def xlsx_response(workbook_bytes, filename):
    response = HttpResponse(workbook_bytes,
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def workbook_to_bytes(wb):
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def format_money_cell(cell):
    cell.number_format = '#,##0.00'


def add_common_header(ws, title):
    # School name row
    ws.merge_cells('A1:G1')
    ws['A1'] = getattr(settings, 'SCHOOL_NAME', 'PCEA Wendani Academy')
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Title row
    ws.merge_cells('A2:G2')
    ws['A2'] = title
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    # small meta row
    ws.merge_cells('A3:G3')
    ws['A3'] = f'Generated on: {datetime.now().strftime("%d %b %Y %H:%M")}'
    ws['A3'].font = Font(size=9)
    ws['A3'].alignment = Alignment(horizontal='center')


# ---------- Invoice Summary Report Exports ----------
class InvoiceSummaryReportExcelView(LoginRequiredMixin, View):
    """Exports invoice summary report to Excel (category-based summary)."""

    def get(self, request):
        from .forms import InvoiceSummaryReportFilterForm
        form = InvoiceSummaryReportFilterForm(request.GET)
        form.is_valid()  # Populate cleaned_data

        # Use getattr with empty dict fallback for safety
        cleaned = getattr(form, 'cleaned_data', {})
        academic_year = cleaned.get('academic_year')
        term = cleaned.get('term')
        show_zero = cleaned.get('show_zero_rows', False)

        # If no academic year/term provided, use the most recent
        if not academic_year:
            academic_year = AcademicYear.objects.filter(is_current=True).first() or AcademicYear.objects.order_by('-year').first()
        if not term:
            term = 'term_1'  # Default to term 1
        
        if not academic_year:
            return HttpResponseBadRequest("No academic year found. Please create one first.")

        # Select invoices for the academic year & term
        invoices = Invoice.objects.filter(term__academic_year=academic_year, term__term=term)
        
        # Apply organization filter
        organization = getattr(request, 'organization', None)
        if organization:
            invoices = invoices.filter(organization=organization)
        
        # Only include active students
        invoices = invoices.filter(student__status='active')

        # All invoice items for those invoices
        items_qs = InvoiceItem.objects.filter(invoice__in=invoices, is_active=True)

        # Sum billed per category
        billed_qs = items_qs.values('category').annotate(total_billed=Sum('net_amount'))
        billed_map = {row['category']: (row['total_billed'] or Decimal('0.00')) for row in billed_qs}

        # Collected per category
        collected_map = {}
        if PaymentAllocation is not None:
            alloc_qs = PaymentAllocation.objects.filter(
                invoice_item__in=items_qs,
                is_active=True,
                payment__is_active=True,
                payment__status='completed'
            ).values('invoice_item__category').annotate(collected=Sum('amount'))
            collected_map = {row['invoice_item__category']: (row['collected'] or Decimal('0.00')) for row in alloc_qs}
        else:
            collected_map = {}
            for inv in invoices:
                inv_items = inv.items.filter(is_active=True)
                inv_total = sum((i.net_amount or Decimal('0.00')) for i in inv_items)
                paid = inv.amount_paid or Decimal('0.00')
                if inv_total <= Decimal('0.00'):
                    continue
                for it in inv_items:
                    cat = it.category
                    share = ((it.net_amount or Decimal('0.00')) / inv_total) * paid
                    collected_map[cat] = collected_map.get(cat, Decimal('0.00')) + (share or Decimal('0.00'))

        # Build the set of categories present
        categories = set(billed_map.keys()) | set(collected_map.keys())
        preferred_order = ['tuition', 'meals', 'assessment', 'activity', 'transport', 'other']
        ordered = []
        for p in preferred_order:
            if p in categories:
                ordered.append(p)
        for c in sorted(categories):
            if c not in ordered:
                ordered.append(c)

        rows = []
        total_billed = Decimal('0.00')
        total_collected = Decimal('0.00')
        total_outstanding = Decimal('0.00')
        for cat in ordered:
            billed = billed_map.get(cat, Decimal('0.00'))
            collected = collected_map.get(cat, Decimal('0.00'))
            outstanding = billed - collected
            if not show_zero and billed == Decimal('0.00') and collected == Decimal('0.00') and outstanding == Decimal('0.00'):
                continue
            rows.append({
                'category': cat,
                'total_billed': billed,
                'collected': collected,
                'outstanding': outstanding
            })
            total_billed += billed
            total_collected += collected
            total_outstanding += outstanding

        # Calculate balance B/F and prepayment totals from invoices
        balance_bf_total = invoices.aggregate(total=Sum('balance_bf'))['total'] or Decimal('0.00')
        prepayment_total = invoices.aggregate(total=Sum('prepayment'))['total'] or Decimal('0.00')

        # Build workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice Summary"

        add_common_header(ws, f"Invoice Summary Report - {academic_year.year} Term {term}")

        headers = ['Category', 'Total Billed (KES)', 'Collected (KES)', 'Outstanding (KES)', 'Bal B/F (KES)', 'Prepayment (KES)']
        for c, h in enumerate(headers, start=1):
            ws.cell(row=5, column=c, value=h).font = Font(bold=True)

        row_num = 6
        for r in rows:
            ws.cell(row=row_num, column=1, value=r['category'].title())
            ws.cell(row=row_num, column=2, value=float(r['total_billed']))
            format_money_cell(ws.cell(row=row_num, column=2))
            ws.cell(row=row_num, column=3, value=float(r['collected']))
            format_money_cell(ws.cell(row=row_num, column=3))
            ws.cell(row=row_num, column=4, value=float(r['outstanding']))
            format_money_cell(ws.cell(row=row_num, column=4))
            # Balance B/F and Prepayment are invoice-level, so show empty for category rows
            ws.cell(row=row_num, column=5, value='-')
            ws.cell(row=row_num, column=6, value='-')
            row_num += 1

        # Totals row
        ws.cell(row=row_num, column=1, value='TOTALS').font = Font(bold=True)
        ws.cell(row=row_num, column=2, value=float(total_billed))
        format_money_cell(ws.cell(row=row_num, column=2))
        ws.cell(row=row_num, column=3, value=float(total_collected))
        format_money_cell(ws.cell(row=row_num, column=3))
        ws.cell(row=row_num, column=4, value=float(total_outstanding))
        format_money_cell(ws.cell(row=row_num, column=4))
        ws.cell(row=row_num, column=5, value=float(balance_bf_total))
        format_money_cell(ws.cell(row=row_num, column=5))
        ws.cell(row=row_num, column=6, value=float(prepayment_total))
        format_money_cell(ws.cell(row=row_num, column=6))

        # Auto width columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        bytes_data = workbook_to_bytes(wb)
        filename = f"invoice-summary-{academic_year.year}-term{term}-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
        return xlsx_response(bytes_data, filename)


class InvoiceSummaryReportPDFView(LoginRequiredMixin, View):
    """Generates PDF for invoice summary report."""

    def get(self, request):
        from .forms import InvoiceSummaryReportFilterForm
        form = InvoiceSummaryReportFilterForm(request.GET)
        form.is_valid()  # Populate cleaned_data

        # Use getattr with empty dict fallback for safety
        cleaned = getattr(form, 'cleaned_data', {})
        academic_year = cleaned.get('academic_year')
        term = cleaned.get('term')
        show_zero = cleaned.get('show_zero_rows', False)

        # If no academic year/term provided, use the most recent
        if not academic_year:
            academic_year = AcademicYear.objects.filter(is_current=True).first() or AcademicYear.objects.order_by('-year').first()
        if not term:
            term = 'term_1'  # Default to term 1
        
        if not academic_year:
            return HttpResponseBadRequest("No academic year found. Please create one first.")

        # Reuse logic from InvoiceSummaryReportExcelView to get data
        invoices = Invoice.objects.filter(term__academic_year=academic_year, term__term=term)
        
        # Apply organization filter
        organization = getattr(request, 'organization', None)
        if organization:
            invoices = invoices.filter(organization=organization)
        
        # Only include active students
        invoices = invoices.filter(student__status='active')
        
        items_qs = InvoiceItem.objects.filter(invoice__in=invoices, is_active=True)

        billed_qs = items_qs.values('category').annotate(total_billed=Sum('net_amount'))
        billed_map = {row['category']: (row['total_billed'] or Decimal('0.00')) for row in billed_qs}

        collected_map = {}
        if PaymentAllocation is not None:
            alloc_qs = PaymentAllocation.objects.filter(
                invoice_item__in=items_qs,
                is_active=True,
                payment__is_active=True,
                payment__status='completed'
            ).values('invoice_item__category').annotate(collected=Sum('amount'))
            collected_map = {row['invoice_item__category']: (row['collected'] or Decimal('0.00')) for row in alloc_qs}
        else:
            collected_map = {}
            for inv in invoices:
                inv_items = inv.items.filter(is_active=True)
                inv_total = sum((i.net_amount or Decimal('0.00')) for i in inv_items)
                paid = inv.amount_paid or Decimal('0.00')
                if inv_total <= Decimal('0.00'):
                    continue
                for it in inv_items:
                    cat = it.category
                    share = ((it.net_amount or Decimal('0.00')) / inv_total) * paid
                    collected_map[cat] = collected_map.get(cat, Decimal('0.00')) + (share or Decimal('0.00'))

        categories = set(billed_map.keys()) | set(collected_map.keys())
        preferred_order = ['tuition', 'meals', 'assessment', 'activity', 'transport', 'other']
        ordered = []
        for p in preferred_order:
            if p in categories:
                ordered.append(p)
        for c in sorted(categories):
            if c not in ordered:
                ordered.append(c)

        rows = []
        total_billed = Decimal('0.00')
        total_collected = Decimal('0.00')
        total_outstanding = Decimal('0.00')
        for cat in ordered:
            billed = billed_map.get(cat, Decimal('0.00'))
            collected = collected_map.get(cat, Decimal('0.00'))
            outstanding = billed - collected
            if not show_zero and billed == Decimal('0.00') and collected == Decimal('0.00') and outstanding == Decimal('0.00'):
                continue
            rows.append({
                'category': cat,
                'total_billed': billed,
                'collected': collected,
                'outstanding': outstanding
            })
            total_billed += billed
            total_collected += collected
            total_outstanding += outstanding

        # Calculate balance B/F and prepayment totals from invoices
        balance_bf_total = invoices.aggregate(total=Sum('balance_bf'))['total'] or Decimal('0.00')
        prepayment_total = invoices.aggregate(total=Sum('prepayment'))['total'] or Decimal('0.00')

        context = {
            'report_rows': rows,
            'totals': {
                'billed': total_billed,
                'collected': total_collected,
                'outstanding': total_outstanding,
                'balance_bf': balance_bf_total,
                'prepayment': prepayment_total
            },
            'academic_year': academic_year,
            'term': term,
            'SCHOOL_NAME': getattr(settings, 'SCHOOL_NAME', 'PCEA Wendani Academy'),
            'SCHOOL_LOGO_URL': request.build_absolute_uri(settings.STATIC_URL + 'assets/images/logo.jpeg'),
            'SPONSOR_LOGO_URL': request.build_absolute_uri(settings.STATIC_URL + 'assets/images/logo2.jpeg'),
            'SCHOOL_ADDRESS': getattr(settings, 'SCHOOL_ADDRESS', ''),
            'SCHOOL_CONTACT': getattr(settings, 'SCHOOL_CONTACT', ''),
            'BANK_DETAILS': getattr(settings, 'SCHOOL_BANK_DETAILS', {
                'equity': {'name': 'EQUITY BANK', 'account_name': 'P.C.E.A Wendani Academy',
                           'account_no': '1130280029105'},
                'coop': {'name': 'CO-OPERATIVE BANK', 'account_name': 'P.C.E.A Wendani Academy',
                         'account_no': '01129158350600'},
                'paybills': [
                    {'label': 'PAYBILL (247247)', 'acc_format': '80029#<admission_number>'},
                    {'label': 'PAYBILL (400222)', 'acc_format': '393939#<admission_number>'},
                ]
            }),
            'generated_by': request.user.get_full_name(),
            'generated_on': datetime.now(),
        }

        html_string = render_to_string('reports/pdf/invoice_summary_report_pdf.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
        pdf_bytes = html.write_pdf()
        filename = f"invoice-summary-{academic_year.year}-term{term}-{datetime.now().strftime('%Y%m%d-%H%M')}.pdf"
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ---------- Invoice Detailed Report Exports ----------
class InvoiceDetailedReportExcelView(LoginRequiredMixin, OrganizationFilterMixin, View):
    """Exports invoice detailed report to Excel."""

    def get(self, request):
        from .forms import InvoiceReportFilterForm
        form = InvoiceReportFilterForm(request.GET)
        form.is_valid()

        cleaned = getattr(form, 'cleaned_data', {})
        academic_year = cleaned.get('academic_year')
        term = cleaned.get('term')
        # Read student_class directly from request.GET to ensure we get the value
        student_class = request.GET.get('student_class', '').strip() or cleaned.get('student_class') or ''
        if not student_class:
            student_class = None
        name = cleaned.get('name') or ''
        admission = cleaned.get('admission') or ''
        category = cleaned.get('category') or ''
        start_date = cleaned.get('start_date')
        end_date = cleaned.get('end_date')
        show_all = cleaned.get('show_all', False)

        # Base queryset: ALL invoices (not just other items)
        invoices_qs = Invoice.objects.filter(
            is_active=True
        ).exclude(
            status=InvoiceStatus.CANCELLED
        ).select_related('student', 'term__academic_year')

        # Apply organization filter
        organization = getattr(request, 'organization', None)
        if organization:
            invoices_qs = invoices_qs.filter(organization=organization)

        # Only include active students
        invoices_qs = invoices_qs.filter(student__status='active')

        # Apply filters
        if not show_all:
            if academic_year:
                invoices_qs = invoices_qs.filter(term__academic_year=academic_year)
                if term:
                    invoices_qs = invoices_qs.filter(term__term=term)
            if student_class:
                invoices_qs = invoices_qs.filter(student__current_class__name=student_class)
            if name:
                invoices_qs = invoices_qs.filter(
                    Q(student__first_name__icontains=name) |
                    Q(student__middle_name__icontains=name) |
                    Q(student__last_name__icontains=name)
                )
            if admission:
                invoices_qs = invoices_qs.filter(student__admission_number__icontains=admission)
            if start_date:
                invoices_qs = invoices_qs.filter(issue_date__gte=start_date)
            if end_date:
                invoices_qs = invoices_qs.filter(issue_date__lte=end_date)

        # Group by student to get total invoice amounts
        grouped = invoices_qs.values(
            'student__pk',
            'student__first_name',
            'student__middle_name',
            'student__last_name',
            'student__admission_number',
            'student__current_class__name',
        ).annotate(
            total_billed=Coalesce(Sum('total_amount'), Value(Decimal('0.00')), output_field=DecimalField()),
            total_paid=Coalesce(Sum('amount_paid'), Value(Decimal('0.00')), output_field=DecimalField()),
            total_balance=Coalesce(Sum('balance'), Value(Decimal('0.00')), output_field=DecimalField())
        ).order_by(
            'student__first_name',
            'student__last_name'
        )

        # Get "other" item descriptions for each student (for category column)
        student_pks = [g['student__pk'] for g in grouped]
        other_items_qs = InvoiceItem.objects.filter(
            invoice__student__pk__in=student_pks,
            category='other',
            is_active=True
        )
        
        # Apply category filter if provided
        if category and not show_all:
            other_items_qs = other_items_qs.filter(description__icontains=category)
            # If category filter is applied, only include students who have matching "other" items
            matching_student_pks = set(other_items_qs.values_list('invoice__student__pk', flat=True).distinct())
        
        # Build map of student_pk -> list of descriptions
        other_items_map = {}
        for item in other_items_qs.select_related('invoice__student'):
            student_pk = item.invoice.student.pk
            if student_pk not in other_items_map:
                other_items_map[student_pk] = []
            if item.description:
                other_items_map[student_pk].append(item.description)

        # Build rows
        rows = []
        total_billed = Decimal('0.00')
        total_collected = Decimal('0.00')
        total_balance = Decimal('0.00')

        for g in grouped:
            student_pk = g.get('student__pk')
            
            # If category filter is applied, only include students with matching "other" items
            if category and not show_all:
                if student_pk not in matching_student_pks:
                    continue
            
            # Get category from "other" items description
            other_descriptions = other_items_map.get(student_pk, [])
            category_desc = ', '.join(other_descriptions) if other_descriptions else '—'
            
            first_name = g.get('student__first_name') or ''
            middle_name = g.get('student__middle_name') or ''
            last_name = g.get('student__last_name') or ''
            student_name = f"{first_name} {middle_name} {last_name}".strip()
            student_name = ' '.join(student_name.split())

            admission = g.get('student__admission_number') or ''
            student_cls = g.get('student__current_class__name') or 'Not assigned'
            billed = Decimal(g.get('total_billed') or 0)
            paid = Decimal(g.get('total_paid') or 0)
            balance = Decimal(g.get('total_balance') or 0)

            rows.append({
                'student_name': student_name,
                'admission': admission,
                'student_class': student_cls,
                'category': category_desc,
                'billed': billed,
                'collected': paid,
                'balance': balance
            })

            total_billed += billed
            total_collected += paid
            total_balance += balance

        # Build workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice Report"

        title = "Invoice Report"
        if academic_year:
            title += f" - {academic_year.year}"
            if term:
                title += f" Term {term}"
        add_common_header(ws, title)

        headers = ['Sr no.', 'Student Name', 'Admission Number', 'Grade', 'Category', 'Total Amount', 'Paid', 'Balance']
        for c, h in enumerate(headers, start=1):
            ws.cell(row=5, column=c, value=h).font = Font(bold=True)

        row_num = 6
        for idx, r in enumerate(rows, start=1):
            ws.cell(row=row_num, column=1, value=idx)
            ws.cell(row=row_num, column=2, value=r['student_name'])
            ws.cell(row=row_num, column=3, value=r['admission'])
            ws.cell(row=row_num, column=4, value=r['student_class'])
            ws.cell(row=row_num, column=5, value=r['category'])

            # Money columns
            c6 = ws.cell(row=row_num, column=6, value=float(r['billed']))
            format_money_cell(c6)
            c7 = ws.cell(row=row_num, column=7, value=float(r['collected']))
            format_money_cell(c7)
            c8 = ws.cell(row=row_num, column=8, value=float(r['balance']))
            format_money_cell(c8)

            row_num += 1

        # Totals
        ws.cell(row=row_num, column=1, value='TOTALS').font = Font(bold=True)
        ws.cell(row=row_num, column=5, value='TOTALS').font = Font(bold=True)
        c6 = ws.cell(row=row_num, column=6, value=float(total_billed))
        format_money_cell(c6)
        c6.font = Font(bold=True)
        c7 = ws.cell(row=row_num, column=7, value=float(total_collected))
        format_money_cell(c7)
        c7.font = Font(bold=True)
        c8 = ws.cell(row=row_num, column=8, value=float(total_balance))
        format_money_cell(c8)
        c8.font = Font(bold=True)

        # Auto width columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        bytes_data = workbook_to_bytes(wb)
        filename = f"invoice-report-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
        return xlsx_response(bytes_data, filename)


class InvoiceDetailedReportPDFView(LoginRequiredMixin, OrganizationFilterMixin, View):
    """Generates PDF for invoice detailed report."""

    def get(self, request):
        from .forms import InvoiceDetailedReportFilterForm
        form = InvoiceDetailedReportFilterForm(request.GET)
        form.is_valid()

        cleaned = getattr(form, 'cleaned_data', {})
        academic_year = cleaned.get('academic_year')
        term = cleaned.get('term')
        # Read student_class directly from request.GET to ensure we get the value
        student_class = request.GET.get('student_class', '').strip() or cleaned.get('student_class') or ''
        if not student_class:
            student_class = None
        name = cleaned.get('name') or ''
        admission = cleaned.get('admission') or ''
        category = cleaned.get('category') or ''
        start_date = cleaned.get('start_date')
        end_date = cleaned.get('end_date')
        show_all = cleaned.get('show_all', False)

        # Base queryset: ALL invoices (not just other items)
        invoices_qs = Invoice.objects.filter(
            is_active=True
        ).exclude(
            status=InvoiceStatus.CANCELLED
        ).select_related('student', 'term__academic_year')

        # Apply organization filter
        organization = getattr(request, 'organization', None)
        if organization:
            invoices_qs = invoices_qs.filter(organization=organization)

        # Only include active students
        invoices_qs = invoices_qs.filter(student__status='active')

        # Apply filters (same logic as Excel view)
        if not show_all:
            if academic_year:
                invoices_qs = invoices_qs.filter(term__academic_year=academic_year)
                if term:
                    invoices_qs = invoices_qs.filter(term__term=term)
            if student_class:
                invoices_qs = invoices_qs.filter(student__current_class__name=student_class)
            if name:
                invoices_qs = invoices_qs.filter(
                    Q(student__first_name__icontains=name) |
                    Q(student__middle_name__icontains=name) |
                    Q(student__last_name__icontains=name)
                )
            if admission:
                invoices_qs = invoices_qs.filter(student__admission_number__icontains=admission)
            if start_date:
                invoices_qs = invoices_qs.filter(issue_date__gte=start_date)
            if end_date:
                invoices_qs = invoices_qs.filter(issue_date__lte=end_date)

        # Group by student to get total invoice amounts
        grouped = invoices_qs.values(
            'student__pk',
            'student__first_name',
            'student__middle_name',
            'student__last_name',
            'student__admission_number',
            'student__current_class__name',
        ).annotate(
            total_billed=Coalesce(Sum('total_amount'), Value(Decimal('0.00')), output_field=DecimalField()),
            total_paid=Coalesce(Sum('amount_paid'), Value(Decimal('0.00')), output_field=DecimalField()),
            total_balance=Coalesce(Sum('balance'), Value(Decimal('0.00')), output_field=DecimalField())
        ).order_by(
            'student__first_name',
            'student__last_name'
        )

        # Get "other" item descriptions for each student (for category column)
        student_pks = [g['student__pk'] for g in grouped]
        other_items_qs = InvoiceItem.objects.filter(
            invoice__student__pk__in=student_pks,
            category='other',
            is_active=True
        )
        
        # Apply category filter if provided
        if category and not show_all:
            other_items_qs = other_items_qs.filter(description__icontains=category)
            # If category filter is applied, only include students who have matching "other" items
            matching_student_pks = set(other_items_qs.values_list('invoice__student__pk', flat=True).distinct())
        
        # Build map of student_pk -> list of descriptions
        other_items_map = {}
        for item in other_items_qs.select_related('invoice__student'):
            student_pk = item.invoice.student.pk
            if student_pk not in other_items_map:
                other_items_map[student_pk] = []
            if item.description:
                other_items_map[student_pk].append(item.description)

        rows = []
        total_billed = Decimal('0.00')
        total_collected = Decimal('0.00')
        total_balance = Decimal('0.00')

        for g in grouped:
            student_pk = g.get('student__pk')
            
            # If category filter is applied, only include students with matching "other" items
            if category and not show_all:
                if student_pk not in matching_student_pks:
                    continue
            
            # Get category from "other" items description
            other_descriptions = other_items_map.get(student_pk, [])
            category_desc = ', '.join(other_descriptions) if other_descriptions else '—'
            
            first_name = g.get('student__first_name') or ''
            middle_name = g.get('student__middle_name') or ''
            last_name = g.get('student__last_name') or ''
            student_name = f"{first_name} {middle_name} {last_name}".strip()
            student_name = ' '.join(student_name.split())

            admission = g.get('student__admission_number') or ''
            student_cls = g.get('student__current_class__name') or 'Not assigned'
            billed = Decimal(g.get('total_billed') or 0)
            paid = Decimal(g.get('total_paid') or 0)
            balance = Decimal(g.get('total_balance') or 0)

            rows.append({
                'student_name': student_name,
                'admission': admission,
                'student_class': student_cls,
                'category': category_desc,
                'billed': billed,
                'collected': paid,
                'balance': balance
            })

            total_billed += billed
            total_collected += paid
            total_balance += balance

        context = {
            'rows': rows,
            'totals': {
                'total_billed': total_billed,
                'total_collected': total_collected,
                'total_balance': total_balance,
            },
            'filters': {
                'academic_year': academic_year,
                'term': term,
                'student_class': student_class,
                'name': name,
                'admission': admission,
                'category': category,
                'start_date': start_date,
                'end_date': end_date,
            },
            'SCHOOL_NAME': getattr(settings, 'SCHOOL_NAME', 'PCEA Wendani Academy'),
            'SCHOOL_LOGO_URL': request.build_absolute_uri(settings.STATIC_URL + 'assets/images/logo.jpeg'),
            'SPONSOR_LOGO_URL': request.build_absolute_uri(settings.STATIC_URL + 'assets/images/logo2.jpeg'),
            'SCHOOL_ADDRESS': getattr(settings, 'SCHOOL_ADDRESS', ''),
            'SCHOOL_CONTACT': getattr(settings, 'SCHOOL_CONTACT', ''),
            'generated_by': request.user.get_full_name(),
            'generated_on': datetime.now(),
        }

        html_string = render_to_string('reports/pdf/invoice_detailed_report_pdf.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
        pdf_bytes = html.write_pdf()
        filename = f"invoice-report-{datetime.now().strftime('%Y%m%d-%H%M')}.pdf"
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ---------- Fees Collection Exports ----------
class FeesCollectionExcelView(LoginRequiredMixin, View):
    """Exports fees collection report to Excel."""

    def get(self, request):
        from .forms import FeesCollectionFilterForm
        from datetime import datetime as dt
        from django.utils import timezone

        # Always pass request.GET (even if empty) to create a bound form
        form = FeesCollectionFilterForm(request.GET)
        form.is_valid()  # Populate cleaned_data

        # Extract filters - use getattr with empty dict fallback for safety
        cleaned = getattr(form, 'cleaned_data', {})
        start_date = cleaned.get('start_date')
        end_date = cleaned.get('end_date')
        payment_source = cleaned.get('payment_source') or ''
        selected_class = cleaned.get('student_class') or ''
        group_by = cleaned.get('group_by') or 'none'

        # Base queryset
        payments_qs = Payment.objects.all()

        # Fix datetime warnings - convert dates to timezone-aware datetimes
        if start_date:
            start_datetime = timezone.make_aware(dt.combine(start_date, dt.min.time()))
            payments_qs = payments_qs.filter(payment_date__gte=start_datetime)
        if end_date:
            end_datetime = timezone.make_aware(dt.combine(end_date, dt.max.time()))
            payments_qs = payments_qs.filter(payment_date__lte=end_datetime)

        # Filter by payment source
        if payment_source:
            payments_qs = payments_qs.filter(payment_source=payment_source)

        if selected_class:
            payments_qs = payments_qs.filter(
                Q(student__current_class=selected_class) |
                Q(invoice__student__current_class=selected_class)
            )

        # Build workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Collections"

        title = "Fees Collection Report"
        if start_date and end_date:
            title += f" ({start_date} to {end_date})"
        add_common_header(ws, title)

        headers = ['Payment Date', 'Receipt/Ref', 'Student', 'Admission #', 'Class', 'Amount (KES)', 'Method',
                   'Bank/Source']
        for c, h in enumerate(headers, start=1):
            ws.cell(row=5, column=c, value=h).font = Font(bold=True)

        row_num = 6
        total_amount = Decimal('0.00')

        payments_list = payments_qs.select_related('student', 'invoice').order_by('payment_date')
        for p in payments_list:
            # Get student details
            student_name = None
            student_class_obj = None  # Keep as object first
            admission = None

            if hasattr(p, 'student') and p.student:
                student_name = getattr(p.student, 'full_name', '')
                student_class_obj = getattr(p.student, 'current_class', '')
                admission = getattr(p.student, 'admission_number', '')
            elif hasattr(p, 'invoice') and getattr(p, 'invoice', None) and getattr(p.invoice, 'student', None):
                st = p.invoice.student
                student_name = getattr(st, 'full_name', '')
                student_class_obj = getattr(st, 'current_class', '')
                admission = getattr(st, 'admission_number', '')
            else:
                student_name = getattr(p, 'payer_name', '') or getattr(p, 'payment_source', '') or '—'

            # Convert Class object to string
            student_class = str(student_class_obj) if student_class_obj else ''

            bank_display = getattr(p, 'bank', '') or getattr(p, 'payment_source', '') or getattr(p, 'payment_method', '')

            ws.cell(row=row_num, column=1, value=p.payment_date.strftime('%Y-%m-%d %H:%M') if p.payment_date else '')
            ws.cell(row=row_num, column=2, value=getattr(p, 'receipt_number', getattr(p, 'payment_reference', '')))
            ws.cell(row=row_num, column=3, value=student_name)
            ws.cell(row=row_num, column=4, value=admission or '')
            ws.cell(row=row_num, column=5, value=student_class or '')  # Now a string
            amount_cell = ws.cell(row=row_num, column=6, value=float(p.amount or 0))
            format_money_cell(amount_cell)
            ws.cell(row=row_num, column=7, value=p.get_payment_method_display() if hasattr(p, 'get_payment_method_display') else getattr(p, 'payment_method', ''))
            ws.cell(row=row_num, column=8, value=bank_display)

            total_amount += Decimal(p.amount or 0)
            row_num += 1

        # Totals row
        ws.cell(row=row_num, column=5, value='TOTAL').font = Font(bold=True)
        total_cell = ws.cell(row=row_num, column=6, value=float(total_amount))
        format_money_cell(total_cell)

        # Auto width columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        bytes_data = workbook_to_bytes(wb)
        filename = f"fees-collections-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
        return xlsx_response(bytes_data, filename)

class FeesCollectionPDFView(LoginRequiredMixin, View):
    """Generates PDF for fees collection report."""

    def get(self, request):
        from .forms import FeesCollectionFilterForm

        # Always pass request.GET (even if empty) to create a bound form
        form = FeesCollectionFilterForm(request.GET)
        form.is_valid()  # Populate cleaned_data

        # Extract filters - use getattr with empty dict fallback for safety
        cleaned = getattr(form, 'cleaned_data', {})
        start_date = cleaned.get('start_date')
        end_date = cleaned.get('end_date')
        payment_source = cleaned.get('payment_source') or ''
        selected_class = cleaned.get('student_class') or ''
        group_by = cleaned.get('group_by') or 'none'

        # Base queryset
        payments_qs = Payment.objects.all()

        if start_date:
            payments_qs = payments_qs.filter(payment_date__gte=start_date)
        if end_date:
            payments_qs = payments_qs.filter(payment_date__lte=end_date)

        # Filter by payment source
        if payment_source:
            payments_qs = payments_qs.filter(payment_source=payment_source)

        if selected_class:
            payments_qs = payments_qs.filter(
                Q(student__current_class=selected_class) |
                Q(invoice__student__current_class=selected_class)
            )

        # Build rows for template
        rows = []
        total_amount = Decimal('0.00')

        payments_list = payments_qs.select_related('student', 'invoice').order_by('payment_date')
        for p in payments_list:
            # Get student details
            student_name = None
            student_class = None
            admission = None

            if hasattr(p, 'student') and p.student:
                student_name = getattr(p.student, 'full_name', '')
                student_class_obj = getattr(p.student, 'current_class', None)
                admission = getattr(p.student, 'admission_number', '')
            elif hasattr(p, 'invoice') and getattr(p, 'invoice', None) and getattr(p.invoice, 'student', None):
                st = p.invoice.student
                student_name = getattr(st, 'full_name', '')
                student_class_obj = getattr(st, 'current_class', None)
                admission = getattr(st, 'admission_number', '')
            else:
                student_name = getattr(p, 'payer_name', '') or getattr(p, 'payment_source', '') or '—'
                student_class_obj = None

            # Convert Class object to string (like in student_list template)
            student_class = str(student_class_obj) if student_class_obj else ''

            bank_display = getattr(p, 'bank', '') or getattr(p, 'payment_source', '') or getattr(p, 'payment_method',
                                                                                                 '')

            rows.append({
                'date': p.payment_date,
                'reference': getattr(p, 'receipt_number', getattr(p, 'payment_reference', '')),
                'student': student_name,
                'class': student_class,
                'admission': admission or '',
                'amount': p.amount or Decimal('0.00'),
                'method': p.get_payment_method_display() if hasattr(p, 'get_payment_method_display') else getattr(p, 'payment_method', ''),
                'bank': bank_display,
            })
            total_amount += Decimal(p.amount or 0)

        context = {
            'rows': rows,
            'summary': {
                'total_collected': total_amount,
                'count': len(rows)
            },
            'filters': {
                'start_date': start_date,
                'end_date': end_date,
                'payment_source': payment_source,
                'student_class': selected_class,
                'group_by': group_by,
            },
            'SCHOOL_NAME': getattr(settings, 'SCHOOL_NAME', 'PCEA Wendani Academy'),
            'SCHOOL_LOGO_URL': request.build_absolute_uri(settings.STATIC_URL + 'assets/images/logo.jpeg'),
            'SPONSOR_LOGO_URL': request.build_absolute_uri(settings.STATIC_URL + 'assets/images/logo2.jpeg'),
            'SCHOOL_ADDRESS': getattr(settings, 'SCHOOL_ADDRESS', ''),
            'SCHOOL_CONTACT': getattr(settings, 'SCHOOL_CONTACT', ''),
            'BANK_DETAILS': getattr(settings, 'SCHOOL_BANK_DETAILS', {
                'equity': {'name': 'EQUITY BANK', 'account_name': 'P.C.E.A Wendani Academy',
                           'account_no': '1130280029105'},
                'coop': {'name': 'CO-OPERATIVE BANK', 'account_name': 'P.C.E.A Wendani Academy',
                         'account_no': '01129158350600'},
                'paybills': [
                    {'label': 'PAYBILL (247247)', 'acc_format': '80029#<admission_number>'},
                    {'label': 'PAYBILL (400222)', 'acc_format': '393939#<admission_number>'},
                ]
            }),
            'generated_on': datetime.now(),
        }

        html_string = render_to_string('reports/pdf/fees_collection_report_pdf.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
        pdf_bytes = html.write_pdf()
        filename = f"fees-collections-{datetime.now().strftime('%Y%m%d-%H%M')}.pdf"
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ---------- Outstanding Balances Exports ----------
class OutstandingBalancesExcelView(LoginRequiredMixin, OrganizationFilterMixin, View):
    """Exports outstanding balances report to Excel."""

    def get(self, request):
        from .forms import OutstandingBalancesFilterForm
        form = OutstandingBalancesFilterForm(request.GET)
        form.is_valid()  # Populate cleaned_data - all fields are optional

        # extract filters - use getattr with empty dict fallback for safety
        cleaned = getattr(form, 'cleaned_data', {})
        start_date = cleaned.get('start_date')
        end_date = cleaned.get('end_date')
        academic_year = cleaned.get('academic_year')
        term = cleaned.get('term')
        # Read student_class directly from request.GET to ensure we get the value even if not in form choices
        student_class = request.GET.get('student_class', '').strip() or cleaned.get('student_class')
        # Ensure student_class is not empty string
        if not student_class:
            student_class = None
        balance_op = cleaned.get('balance_operator') or 'any'
        balance_amt = cleaned.get('balance_amount') or Decimal('0.00')
        include_zero = cleaned.get('show_zero_balances')

        invoices = Invoice.objects.select_related('student', 'term__academic_year')
        
        # Apply organization filter
        organization = getattr(request, 'organization', None)
        if organization:
            invoices = invoices.filter(organization=organization)

        # Only include active students - exclude all non-active statuses
        invoices = invoices.filter(student__status='active')

        if academic_year:
            invoices = invoices.filter(term__academic_year=academic_year)
            if term:
                invoices = invoices.filter(term__term=term)
        if start_date:
            invoices = invoices.filter(issue_date__gte=start_date)
        if end_date:
            invoices = invoices.filter(issue_date__lte=end_date)
        # Apply student_class filter - ensure it's applied correctly
        if student_class:
            invoices = invoices.filter(student__current_class__name=student_class)

        annotations = {
            'total_billed': ExpressionWrapper(
                Coalesce(Sum('total_amount'), Value(Decimal('0.00'))),
                output_field=DecimalField()
            ),
            'total_paid': ExpressionWrapper(
                Coalesce(Sum('amount_paid'), Value(Decimal('0.00'))),
                output_field=DecimalField()
            ),
            'total_balance': ExpressionWrapper(
                Coalesce(Sum('balance'), Value(Decimal('0.00'))),
                output_field=DecimalField()
            ),
            'total_balance_bf': ExpressionWrapper(
                Coalesce(Sum('balance_bf'), Value(Decimal('0.00'))),
                output_field=DecimalField()
            ),
            'total_prepayment': ExpressionWrapper(
                Coalesce(Sum('prepayment'), Value(Decimal('0.00'))),
                output_field=DecimalField()
            ),
        }

        grouped_qs = invoices.values(
            'student__pk',
            'student__admission_number',
            'student__first_name',
            'student__middle_name',
            'student__last_name',
            'student__current_class__name',
            'term__academic_year__year',
        ).annotate(**annotations).order_by('-total_balance', 'student__first_name', 'student__last_name')

        # balance filter
        if balance_op != 'any' and balance_amt is not None:
            lookup = {
                '=': 'total_balance',
                '>': 'total_balance__gt',
                '<': 'total_balance__lt',
                '>=': 'total_balance__gte',
                '<=': 'total_balance__lte',
            }.get(balance_op, None)
            if lookup:
                grouped_qs = grouped_qs.filter(**{lookup: balance_amt})

        if not include_zero:
            grouped_qs = grouped_qs.exclude(total_balance=Decimal('0.00'))

        rows = list(grouped_qs)

        # Build Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Outstanding Balances"

        title = "Outstanding Balances Report"
        if academic_year:
            title += f" - {academic_year.year}"
            if term:
                title += f" Term {term}"
        add_common_header(ws, title)

        headers = ['Year', 'Admission No', 'Name', 'Class', 'Emergency Contact', 'Balance B/F', 'Prepayment',
                   'Total Billed',
                   'Paid', 'Balance']
        for c, h in enumerate(headers, start=1):
            ws.cell(row=5, column=c, value=h).font = Font(bold=True)

        row_num = 6
        for r in rows:
            ws.cell(row=row_num, column=1, value=r.get('term__academic_year__year'))
            ws.cell(row=row_num, column=2, value=r.get('student__admission_number'))

            # Build full name from first, middle, last
            first = r.get('student__first_name', '')
            middle = r.get('student__middle_name', '')
            last = r.get('student__last_name', '')
            full_name = f"{first} {middle} {last}".strip()
            full_name = ' '.join(full_name.split())
            ws.cell(row=row_num, column=3, value=full_name)
            
            # Class
            ws.cell(row=row_num, column=4, value=r.get('student__current_class__name') or '—')
            
            # Emergency Contact (placeholder for now)
            ws.cell(row=row_num, column=5, value='—')

            # Money columns
            money_columns = [6, 7, 8, 9, 10]
            values = [
                float(r.get('total_balance_bf') or 0),
                float(r.get('total_prepayment') or 0),
                float(r.get('total_billed') or 0),
                float(r.get('total_paid') or 0),
                float(r.get('total_balance') or 0)
            ]

            for col_idx, value in zip(money_columns, values):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                format_money_cell(cell)

            row_num += 1

        # Totals row
        totals = {
            'total_balance_bf': sum((r['total_balance_bf'] or Decimal('0.00')) for r in rows),
            'total_prepayment': sum((r['total_prepayment'] or Decimal('0.00')) for r in rows),
            'total_billed': sum((r['total_billed'] or Decimal('0.00')) for r in rows),
            'total_paid': sum((r['total_paid'] or Decimal('0.00')) for r in rows),
            'total_balance': sum((r['total_balance'] or Decimal('0.00')) for r in rows),
        }

        ws.cell(row=row_num, column=1, value='TOTALS').font = Font(bold=True)
        for col_idx, value in zip(money_columns, [
            float(totals['total_balance_bf']),
            float(totals['total_prepayment']),
            float(totals['total_billed']),
            float(totals['total_paid']),
            float(totals['total_balance'])
        ]):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            format_money_cell(cell)

        # Auto width columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        bytes_data = workbook_to_bytes(wb)
        filename = f"outstanding-balances-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
        return xlsx_response(bytes_data, filename)


class OutstandingBalancesPDFView(LoginRequiredMixin, OrganizationFilterMixin, View):
    """Generates PDF for outstanding balances report."""

    def get(self, request):
        from .forms import OutstandingBalancesFilterForm
        form = OutstandingBalancesFilterForm(request.GET)
        form.is_valid()  # Populate cleaned_data - all fields are optional

        # Use getattr with empty dict fallback for safety
        cleaned = getattr(form, 'cleaned_data', {})
        start_date = cleaned.get('start_date')
        end_date = cleaned.get('end_date')
        academic_year = cleaned.get('academic_year')
        term = cleaned.get('term')
        # Read student_class directly from request.GET to ensure we get the value even if not in form choices
        student_class = request.GET.get('student_class', '').strip() or cleaned.get('student_class')
        # Ensure student_class is not empty string
        if not student_class:
            student_class = None
        balance_op = cleaned.get('balance_operator') or 'any'
        balance_amt = cleaned.get('balance_amount') or Decimal('0.00')
        include_zero = cleaned.get('show_zero_balances')

        invoices = Invoice.objects.select_related('student', 'term__academic_year')
        
        # Apply organization filter
        organization = getattr(request, 'organization', None)
        if organization:
            invoices = invoices.filter(organization=organization)

        # Only include active students - exclude all non-active statuses
        invoices = invoices.filter(student__status='active')
            
        if academic_year:
            invoices = invoices.filter(term__academic_year=academic_year)
            if term:
                invoices = invoices.filter(term__term=term)
        if start_date:
            invoices = invoices.filter(issue_date__gte=start_date)
        if end_date:
            invoices = invoices.filter(issue_date__lte=end_date)
        # Apply student_class filter - ensure it's applied correctly
        if student_class:
            invoices = invoices.filter(student__current_class__name=student_class)

        annotations = {
            'total_billed': ExpressionWrapper(
                Coalesce(Sum('total_amount'), Value(Decimal('0.00'))),
                output_field=DecimalField()
            ),
            'total_paid': ExpressionWrapper(
                Coalesce(Sum('amount_paid'), Value(Decimal('0.00'))),
                output_field=DecimalField()
            ),
            'total_balance': ExpressionWrapper(
                Coalesce(Sum('balance'), Value(Decimal('0.00'))),
                output_field=DecimalField()
            ),
            'total_balance_bf': ExpressionWrapper(
                Coalesce(Sum('balance_bf'), Value(Decimal('0.00'))),
                output_field=DecimalField()
            ),
            'total_prepayment': ExpressionWrapper(
                Coalesce(Sum('prepayment'), Value(Decimal('0.00'))),
                output_field=DecimalField()
            ),
        }

        grouped_qs = invoices.values(
            'student__pk',
            'student__admission_number',
            'student__first_name',
            'student__middle_name',
            'student__last_name',
            'student__current_class__name',
            'term__academic_year__year',
        ).annotate(**annotations).order_by('-total_balance', 'student__first_name', 'student__last_name')

        if balance_op != 'any' and balance_amt is not None:
            lookup = {
                '=': 'total_balance',
                '>': 'total_balance__gt',
                '<': 'total_balance__lt',
                '>=': 'total_balance__gte',
                '<=': 'total_balance__lte',
            }.get(balance_op, None)
            if lookup:
                grouped_qs = grouped_qs.filter(**{lookup: balance_amt})

        if not include_zero:
            grouped_qs = grouped_qs.exclude(total_balance=Decimal('0.00'))

        rows = list(grouped_qs)
        totals = {
            'total_balance_bf': sum((r['total_balance_bf'] or Decimal('0.00')) for r in rows),
            'total_prepayment': sum((r['total_prepayment'] or Decimal('0.00')) for r in rows),
            'total_billed': sum((r['total_billed'] or Decimal('0.00')) for r in rows),
            'total_paid': sum((r['total_paid'] or Decimal('0.00')) for r in rows),
            'total_balance': sum((r['total_balance'] or Decimal('0.00')) for r in rows),
        }

        # Process rows to include full_name and contact_info
        processed_rows = []
        for r in rows:
            # Build full name
            first = r.get('student__first_name', '')
            middle = r.get('student__middle_name', '')
            last = r.get('student__last_name', '')
            full_name = f"{first} {middle} {last}".strip()
            full_name = ' '.join(full_name.split())



            processed_rows.append({
                'term__academic_year__year': r.get('term__academic_year__year'),
                'student__admission_number': r.get('student__admission_number'),
                'student__first_name': first,
                'student__middle_name': middle,
                'student__last_name': last,
                'student__current_class__name': r.get('student__current_class__name'),
                'total_balance_bf': r.get('total_balance_bf'),
                'total_prepayment': r.get('total_prepayment'),
                'total_billed': r.get('total_billed'),
                'total_paid': r.get('total_paid'),
                'total_balance': r.get('total_balance'),
            })

        context = {
            'rows': grouped_qs,  # Pass the queryset directly, not processed_rows
            'totals': totals,
            'filters': {
                'start_date': start_date,
                'end_date': end_date,
                'academic_year': academic_year,
                'term': term,
                'balance_op': balance_op,
                'balance_amt': balance_amt,
            },
            'SCHOOL_NAME': getattr(settings, 'SCHOOL_NAME', 'PCEA Wendani Academy'),
            'SCHOOL_LOGO_URL': request.build_absolute_uri(settings.STATIC_URL + 'assets/images/logo.jpeg'),
            'SPONSOR_LOGO_URL': request.build_absolute_uri(settings.STATIC_URL + 'assets/images/logo2.jpeg'),
            'SCHOOL_ADDRESS': getattr(settings, 'SCHOOL_ADDRESS', ''),
            'SCHOOL_CONTACT': getattr(settings, 'SCHOOL_CONTACT', ''),
            'BANK_DETAILS': getattr(settings, 'SCHOOL_BANK_DETAILS', {
                'equity': {'name': 'EQUITY BANK', 'account_name': 'P.C.E.A Wendani Academy',
                           'account_no': '1130280029105'},
                'coop': {'name': 'CO-OPERATIVE BANK', 'account_name': 'P.C.E.A Wendani Academy',
                         'account_no': '01129158350600'},
                'paybills': [
                    {'label': 'PAYBILL (247247)', 'acc_format': '80029#<admission_number>'},
                    {'label': 'PAYBILL (400222)', 'acc_format': '393939#<admission_number>'},
                ]
            }),
            'generated_by': request.user.get_full_name(),
            'generated_on': datetime.now(),
        }

        html_string = render_to_string('reports/pdf/outstanding_report_pdf.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
        pdf_bytes = html.write_pdf()
        filename = f"outstanding-balances-{datetime.now().strftime('%Y%m%d-%H%M')}.pdf"
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ---------- Transport Report Exports ----------
class TransportReportExcelView(LoginRequiredMixin, View):
    """Exports transport report to Excel."""

    def get(self, request):
        from .forms import TransportReportFilterForm
        form = TransportReportFilterForm(request.GET)
        form.is_valid()  # Populate cleaned_data

        # Use getattr with empty dict fallback for safety
        cleaned = getattr(form, 'cleaned_data', {})
        academic_year = cleaned.get('academic_year')
        term = cleaned.get('term')
        
        # If no academic year/term provided, use the most recent
        if not academic_year:
            academic_year = AcademicYear.objects.filter(is_current=True).first() or AcademicYear.objects.order_by('-year').first()
        if not term:
            term = 'term_1'  # Default to term 1
        
        if not academic_year:
            return HttpResponseBadRequest("No academic year found. Please create one first.")
        route = cleaned.get('route')
        student_class = cleaned.get('student_class') or ''
        start_date = cleaned.get('start_date')
        end_date = cleaned.get('end_date')
        show_zero = cleaned.get('show_zero_rows', False)

        items_qs = InvoiceItem.objects.filter(
            invoice__term__academic_year=academic_year,
            invoice__term__term=term,
            category='transport'
        ).select_related('invoice__student', 'transport_route')

        if route:
            items_qs = items_qs.filter(transport_route=route)

        if student_class:
            items_qs = items_qs.filter(invoice__student__current_class=student_class)

        if start_date:
            items_qs = items_qs.filter(invoice__issue_date__gte=start_date)
        if end_date:
            items_qs = items_qs.filter(invoice__issue_date__lte=end_date)

        # FIXED: Use individual name fields instead of full_name
        grouped = items_qs.values(
            'invoice__student__pk',
            'invoice__student__first_name',  # Changed from full_name
            'invoice__student__middle_name',  # Added
            'invoice__student__last_name',  # Added
            'invoice__student__admission_number',
            'invoice__student__current_class__name',  # Get class name instead of UUID
            'transport_route__pk',
            'transport_route__name'
        ).annotate(
            total_billed=Coalesce(Sum('net_amount'), Value(Decimal('0.00')), output_field=DecimalField())
        ).order_by(
            'invoice__student__first_name',
            'invoice__student__last_name'
        )

        # Build collected_map
        collected_map = {}
        try:
            alloc_qs = PaymentAllocation.objects.filter(invoice_item__in=items_qs).values(
                'invoice_item__invoice__student__pk',
                'invoice_item__transport_route__pk'
            ).annotate(
                collected=Coalesce(Sum('amount'), Value(Decimal('0.00')), output_field=DecimalField())
            )
            for row in alloc_qs:
                key = (row.get('invoice_item__invoice__student__pk'), row.get('invoice_item__transport_route__pk'))
                collected_map[key] = Decimal(row.get('collected') or 0)
        except Exception:
            collected_map = {}
            invoice_ids = items_qs.values_list('invoice_id', flat=True).distinct()
            invoices = Invoice.objects.filter(id__in=invoice_ids)
            for inv in invoices:
                inv_items = items_qs.filter(invoice=inv)
                inv_total_transport = sum((i.net_amount or Decimal('0.00')) for i in inv_items)
                paid = inv.amount_paid or Decimal('0.00')
                if inv_total_transport <= Decimal('0.00'):
                    for it in inv_items:
                        key = (inv.student.pk, it.transport_route.pk if it.transport_route else None)
                        collected_map[key] = collected_map.get(key, Decimal('0.00'))
                    continue
                for it in inv_items:
                    proportion = (it.net_amount or Decimal('0.00')) / inv_total_transport
                    alloc_amount = (paid * proportion).quantize(Decimal('0.01'))
                    key = (inv.student.pk, it.transport_route.pk if it.transport_route else None)
                    collected_map[key] = collected_map.get(key, Decimal('0.00')) + alloc_amount

        # Build rows
        rows = []
        total_billed = Decimal('0.00')
        total_collected = Decimal('0.00')
        total_balance = Decimal('0.00')

        for g in grouped:
            student_pk = g.get('invoice__student__pk')
            billed = Decimal(g.get('total_billed') or 0)
            route_pk = g.get('transport_route__pk')
            collected = collected_map.get((student_pk, route_pk), Decimal('0.00'))
            balance = billed - collected

            # Build full name from individual fields
            first_name = g.get('invoice__student__first_name') or ''
            middle_name = g.get('invoice__student__middle_name') or ''
            last_name = g.get('invoice__student__last_name') or ''
            student_name = f"{first_name} {middle_name} {last_name}".strip()
            # Clean up extra spaces
            student_name = ' '.join(student_name.split())

            if (not show_zero) and billed == Decimal('0.00') and collected == Decimal('0.00'):
                continue

            rows.append({
                'student_name': student_name,
                'admission': g.get('invoice__student__admission_number') or '',
                'student_class': g.get('invoice__student__current_class__name') or 'Not assigned',
                'route_name': g.get('transport_route__name') or '',
                'billed': billed,
                'collected': collected,
                'balance': balance,
            })
            total_billed += billed
            total_collected += collected
            total_balance += balance

        # Build workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transport Report"

        title = f"Transport Report - {academic_year.year} Term {term}"
        if route:
            title += f" - {route.name}"
        add_common_header(ws, title)

        headers = ['Student Name', 'Admission #', 'Class/Grade', 'Route/Destination', 'Transport Amount', 'Paid',
                   'Balance']
        for c, h in enumerate(headers, start=1):
            ws.cell(row=5, column=c, value=h).font = Font(bold=True)

        row_num = 6
        for r in rows:
            ws.cell(row=row_num, column=1, value=r['student_name'])
            ws.cell(row=row_num, column=2, value=r['admission'])
            ws.cell(row=row_num, column=3, value=r['student_class'])  # Now available
            ws.cell(row=row_num, column=4, value=r['route_name'])

            # Money columns
            c5 = ws.cell(row=row_num, column=5, value=float(r['billed']))
            format_money_cell(c5)
            c6 = ws.cell(row=row_num, column=6, value=float(r['collected']))
            format_money_cell(c6)
            c7 = ws.cell(row=row_num, column=7, value=float(r['balance']))
            format_money_cell(c7)

            row_num += 1

        # Totals
        ws.cell(row=row_num, column=1, value='TOTALS').font = Font(bold=True)
        c5 = ws.cell(row=row_num, column=5, value=float(total_billed))
        format_money_cell(c5)
        c6 = ws.cell(row=row_num, column=6, value=float(total_collected))
        format_money_cell(c6)
        c7 = ws.cell(row=row_num, column=7, value=float(total_balance))
        format_money_cell(c7)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        bytes_data = workbook_to_bytes(wb)
        filename = f"transport-report-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
        return xlsx_response(bytes_data, filename)


class TransportReportPDFView(LoginRequiredMixin, View):
    """Generates PDF for transport report."""

    def get(self, request):
        from .forms import TransportReportFilterForm
        form = TransportReportFilterForm(request.GET)
        form.is_valid()  # Populate cleaned_data

        # Use getattr with empty dict fallback for safety
        cleaned = getattr(form, 'cleaned_data', {})
        academic_year = cleaned.get('academic_year')
        term = cleaned.get('term')
        
        # If no academic year/term provided, use the most recent
        if not academic_year:
            academic_year = AcademicYear.objects.filter(is_current=True).first() or AcademicYear.objects.order_by('-year').first()
        if not term:
            term = 'term_1'  # Default to term 1
        
        if not academic_year:
            return HttpResponseBadRequest("No academic year found. Please create one first.")
        route = cleaned.get('route')
        student_class = cleaned.get('student_class') or ''
        start_date = cleaned.get('start_date')
        end_date = cleaned.get('end_date')
        show_zero = cleaned.get('show_zero_rows', False)

        items_qs = InvoiceItem.objects.filter(
            invoice__term__academic_year=academic_year,
            invoice__term__term=term,
            category='transport'
        ).select_related('invoice__student', 'transport_route')

        if route:
            items_qs = items_qs.filter(transport_route=route)

        if student_class:
            items_qs = items_qs.filter(invoice__student__current_class=student_class)

        if start_date:
            items_qs = items_qs.filter(invoice__issue_date__gte=start_date)
        if end_date:
            items_qs = items_qs.filter(invoice__issue_date__lte=end_date)

        # FIXED: Use individual name fields instead of full_name and add output_field
        grouped = items_qs.values(
            'invoice__student__pk',
            'invoice__student__first_name',  # Changed from full_name
            'invoice__student__middle_name',  # Added
            'invoice__student__last_name',  # Added
            'invoice__student__admission_number',
            'invoice__student__current_class__name',  # Get class name instead of UUID
            'transport_route__pk',
            'transport_route__name'
        ).annotate(
            total_billed=Coalesce(Sum('net_amount'), Value(Decimal('0.00')), output_field=DecimalField())
        ).order_by(
            'invoice__student__first_name',
            'invoice__student__last_name'
        )

        # Build collected_map
        collected_map = {}
        try:
            alloc_qs = PaymentAllocation.objects.filter(invoice_item__in=items_qs).values(
                'invoice_item__invoice__student__pk',
                'invoice_item__transport_route__pk'
            ).annotate(
                collected=Coalesce(Sum('amount'), Value(Decimal('0.00')), output_field=DecimalField())
            )
            for row in alloc_qs:
                key = (row.get('invoice_item__invoice__student__pk'), row.get('invoice_item__transport_route__pk'))
                collected_map[key] = Decimal(row.get('collected') or 0)
        except Exception:
            collected_map = {}
            invoice_ids = items_qs.values_list('invoice_id', flat=True).distinct()
            invoices = Invoice.objects.filter(id__in=invoice_ids)
            for inv in invoices:
                inv_items = items_qs.filter(invoice=inv)
                inv_total_transport = sum((i.net_amount or Decimal('0.00')) for i in inv_items)
                paid = inv.amount_paid or Decimal('0.00')
                if inv_total_transport <= Decimal('0.00'):
                    for it in inv_items:
                        key = (inv.student.pk, it.transport_route.pk if it.transport_route else None)
                        collected_map[key] = collected_map.get(key, Decimal('0.00'))
                    continue
                for it in inv_items:
                    proportion = (it.net_amount or Decimal('0.00')) / inv_total_transport
                    alloc_amount = (paid * proportion).quantize(Decimal('0.01'))
                    key = (inv.student.pk, it.transport_route.pk if it.transport_route else None)
                    collected_map[key] = collected_map.get(key, Decimal('0.00')) + alloc_amount

        rows = []
        total_billed = Decimal('0.00')
        total_collected = Decimal('0.00')
        total_balance = Decimal('0.00')

        for g in grouped:
            student_pk = g.get('invoice__student__pk')
            billed = Decimal(g.get('total_billed') or 0)
            route_pk = g.get('transport_route__pk')
            collected = collected_map.get((student_pk, route_pk), Decimal('0.00'))
            balance = billed - collected

            # Build full name from individual fields
            first_name = g.get('invoice__student__first_name') or ''
            middle_name = g.get('invoice__student__middle_name') or ''
            last_name = g.get('invoice__student__last_name') or ''
            student_name = f"{first_name} {middle_name} {last_name}".strip()
            # Clean up extra spaces
            student_name = ' '.join(student_name.split())

            if (not show_zero) and billed == Decimal('0.00') and collected == Decimal('0.00'):
                continue

            rows.append({
                'student_name': student_name,
                'admission': g.get('invoice__student__admission_number') or '',
                'student_class': g.get('invoice__student__current_class__name') or 'Not assigned',
                'route_name': g.get('transport_route__name') or '',
                'billed': billed,
                'collected': collected,
                'balance': balance
            })
            total_billed += billed
            total_collected += collected
            total_balance += balance

        context = {
            'rows': rows,
            'totals': {
                'billed': total_billed,
                'collected': total_collected,
                'balance': total_balance
            },
            'filters': {
                'academic_year': academic_year,
                'term': term,
                'route': route,
                'student_class': student_class,
                'start_date': start_date,
                'end_date': end_date,
                'show_zero': show_zero,
            },
            'SCHOOL_NAME': getattr(settings, 'SCHOOL_NAME', 'PCEA Wendani Academy'),
            'SCHOOL_ADDRESS': getattr(settings, 'SCHOOL_ADDRESS', ''),
            'SCHOOL_CONTACT': getattr(settings, 'SCHOOL_CONTACT', ''),
            'generated_by': request.user.get_full_name(),
            'generated_on': datetime.now(),
        }

        html_string = render_to_string('reports/pdf/transport_report_pdf.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
        pdf_bytes = html.write_pdf()
        filename = f"transport-report-{datetime.now().strftime('%Y%m%d-%H%M')}.pdf"
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ---------- Invoice List Exports ----------
class InvoiceListExcelView(LoginRequiredMixin, View):
    """Exports invoice list to Excel with same filters as InvoiceListView."""

    def get_queryset(self):
        """Apply same filtering logic as InvoiceListView."""
        queryset = Invoice.objects.filter(
            is_active=True
        ).select_related('student', 'term', 'term__academic_year')

        query = self.request.GET.get('query', '')
        if query:
            queryset = queryset.filter(
                Q(invoice_number__icontains=query) |
                Q(student__admission_number__icontains=query) |
                Q(student__first_name__icontains=query) |
                Q(student__last_name__icontains=query)
            )

        term = self.request.GET.get('term')
        if term:
            queryset = queryset.filter(term_id=term)

        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        grade = self.request.GET.get('grade')
        if grade:
            queryset = queryset.filter(student__current_class__grade_level=grade)

        return queryset.order_by('-issue_date', '-created_at')

    def get(self, request):
        invoices = self.get_queryset()

        # Build workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice List"

        add_common_header(ws, "Invoice List")

        headers = ['Invoice #', 'Issue Date', 'Student Name', 'Admission No', 'Term', 
                   'Bal B/F (KES)', 'Prepayment (KES)', 'Billed (KES)', 'Discount (KES)', 
                   'Total (KES)', 'Paid (KES)', 'Balance (KES)', 'Status']
        for c, h in enumerate(headers, start=1):
            ws.cell(row=5, column=c, value=h).font = Font(bold=True)

        row_num = 6
        for invoice in invoices:
            student_name = invoice.student.full_name if invoice.student else '—'
            admission_no = invoice.student.admission_number if invoice.student else '—'
            term_str = str(invoice.term) if invoice.term else '—'
            # Get status display - Django automatically provides get_status_display() for CharField with choices
            status_str = invoice.get_status_display() if hasattr(invoice, 'get_status_display') else str(invoice.status)

            ws.cell(row=row_num, column=1, value=invoice.invoice_number)
            ws.cell(row=row_num, column=2, value=invoice.issue_date.strftime('%Y-%m-%d') if invoice.issue_date else '—')
            ws.cell(row=row_num, column=3, value=student_name)
            ws.cell(row=row_num, column=4, value=admission_no)
            ws.cell(row=row_num, column=5, value=term_str)
            ws.cell(row=row_num, column=6, value=float(invoice.balance_bf or 0))
            format_money_cell(ws.cell(row=row_num, column=6))
            ws.cell(row=row_num, column=7, value=float(invoice.prepayment or 0))
            format_money_cell(ws.cell(row=row_num, column=7))
            ws.cell(row=row_num, column=8, value=float(invoice.subtotal or invoice.total_amount or 0))
            format_money_cell(ws.cell(row=row_num, column=8))
            ws.cell(row=row_num, column=9, value=float(invoice.discount_amount or 0))
            format_money_cell(ws.cell(row=row_num, column=9))
            ws.cell(row=row_num, column=10, value=float(invoice.total_amount or 0))
            format_money_cell(ws.cell(row=row_num, column=10))
            ws.cell(row=row_num, column=11, value=float(invoice.amount_paid or 0))
            format_money_cell(ws.cell(row=row_num, column=11))
            ws.cell(row=row_num, column=12, value=float(invoice.balance or 0))
            format_money_cell(ws.cell(row=row_num, column=12))
            ws.cell(row=row_num, column=13, value=status_str)
            row_num += 1

        # Totals row
        totals = invoices.aggregate(
            total_subtotal=Sum('subtotal'),
            total_discount=Sum('discount_amount'),
            total_amount=Sum('total_amount'),
            total_balance_bf=Sum('balance_bf'),
            total_prepayment=Sum('prepayment'),
            total_paid=Sum('amount_paid'),
            total_balance=Sum('balance')
        )
        ws.cell(row=row_num, column=1, value='TOTALS').font = Font(bold=True)
        ws.cell(row=row_num, column=6, value=float(totals['total_balance_bf'] or 0))
        format_money_cell(ws.cell(row=row_num, column=6))
        ws.cell(row=row_num, column=7, value=float(totals['total_prepayment'] or 0))
        format_money_cell(ws.cell(row=row_num, column=7))
        ws.cell(row=row_num, column=8, value=float(totals['total_subtotal'] or totals['total_amount'] or 0))
        format_money_cell(ws.cell(row=row_num, column=8))
        ws.cell(row=row_num, column=9, value=float(totals['total_discount'] or 0))
        format_money_cell(ws.cell(row=row_num, column=9))
        ws.cell(row=row_num, column=10, value=float(totals['total_amount'] or 0))
        format_money_cell(ws.cell(row=row_num, column=10))
        ws.cell(row=row_num, column=11, value=float(totals['total_paid'] or 0))
        format_money_cell(ws.cell(row=row_num, column=11))
        ws.cell(row=row_num, column=12, value=float(totals['total_balance'] or 0))
        format_money_cell(ws.cell(row=row_num, column=12))
        for col in [6, 7, 8, 9, 10, 11, 12]:
            ws.cell(row=row_num, column=col).font = Font(bold=True)

        # Auto width columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        bytes_data = workbook_to_bytes(wb)
        filename = f"invoice-list-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
        return xlsx_response(bytes_data, filename)


class InvoiceListPDFView(LoginRequiredMixin, View):
    """Generates PDF for invoice list."""

    def get_queryset(self):
        """Apply same filtering logic as InvoiceListView."""
        queryset = Invoice.objects.filter(
            is_active=True
        ).select_related('student', 'term', 'term__academic_year')

        query = self.request.GET.get('query', '')
        if query:
            queryset = queryset.filter(
                Q(invoice_number__icontains=query) |
                Q(student__admission_number__icontains=query) |
                Q(student__first_name__icontains=query) |
                Q(student__last_name__icontains=query)
            )

        term = self.request.GET.get('term')
        if term:
            queryset = queryset.filter(term_id=term)

        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        grade = self.request.GET.get('grade')
        if grade:
            queryset = queryset.filter(student__current_class__grade_level=grade)

        return queryset.order_by('-issue_date', '-created_at')

    def get(self, request):
        invoices = self.get_queryset()
        
        # Calculate totals
        totals = invoices.aggregate(
            total_subtotal=Sum('subtotal'),
            total_discount=Sum('discount_amount'),
            total_amount=Sum('total_amount'),
            total_balance_bf=Sum('balance_bf'),
            total_prepayment=Sum('prepayment'),
            total_paid=Sum('amount_paid'),
            total_balance=Sum('balance')
        )

        context = {
            'invoices': invoices,
            'totals': {
                'total_subtotal': totals['total_subtotal'] or Decimal('0.00'),
                'total_discount': totals['total_discount'] or Decimal('0.00'),
                'total_amount': totals['total_amount'] or Decimal('0.00'),
                'total_balance_bf': totals['total_balance_bf'] or Decimal('0.00'),
                'total_prepayment': totals['total_prepayment'] or Decimal('0.00'),
                'total_paid': totals['total_paid'] or Decimal('0.00'),
                'total_balance': totals['total_balance'] or Decimal('0.00'),
            },
            'SCHOOL_NAME': getattr(settings, 'SCHOOL_NAME', 'PCEA Wendani Academy'),
            'SCHOOL_LOGO_URL': request.build_absolute_uri(settings.STATIC_URL + 'assets/images/logo.jpeg'),
            'SCHOOL_ADDRESS': getattr(settings, 'SCHOOL_ADDRESS', ''),
            'SCHOOL_CONTACT': getattr(settings, 'SCHOOL_CONTACT', ''),
            'generated_by': request.user.get_full_name(),
            'generated_on': datetime.now(),
        }

        html_string = render_to_string('reports/pdf/invoice_list_pdf.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
        pdf_bytes = html.write_pdf()
        filename = f"invoice-list-{datetime.now().strftime('%Y%m%d-%H%M')}.pdf"
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ---------- Transferred Students Report Exports ----------
class TransferredStudentsExcelView(LoginRequiredMixin, View):
    """Exports transferred students report to Excel."""

    def get(self, request):
        from .forms import TransferredStudentsFilterForm
        form = TransferredStudentsFilterForm(request.GET)
        form.is_valid()

        cleaned = getattr(form, 'cleaned_data', {})
        academic_year = cleaned.get('academic_year')
        start_date = cleaned.get('start_date')
        end_date = cleaned.get('end_date')
        student_class = cleaned.get('student_class')

        # Base queryset: transferred students
        students_qs = Student.objects.filter(status='transferred').select_related('current_class')

        if academic_year:
            students_qs = students_qs.filter(status_date__year=academic_year.year)
        if start_date:
            students_qs = students_qs.filter(status_date__gte=start_date)
        if end_date:
            students_qs = students_qs.filter(status_date__lte=end_date)
        if student_class:
            students_qs = students_qs.filter(current_class__name=student_class)

        # Build workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transferred Students"

        title = "Transferred Students Report"
        if academic_year:
            title += f" - {academic_year.year}"
        add_common_header(ws, title)

        headers = ['Name', 'Admission Number', 'Grade']
        for c, h in enumerate(headers, start=1):
            ws.cell(row=5, column=c, value=h).font = Font(bold=True)

        row_num = 6
        for student in students_qs.order_by('first_name', 'last_name'):
            ws.cell(row=row_num, column=1, value=student.full_name)
            ws.cell(row=row_num, column=2, value=student.admission_number or 'N/A')
            ws.cell(row=row_num, column=3, value=student.current_class.name if student.current_class else 'Not assigned')
            row_num += 1

        # Auto width columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25

        bytes_data = workbook_to_bytes(wb)
        filename = f"transferred-students-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
        return xlsx_response(bytes_data, filename)


class TransferredStudentsPDFView(LoginRequiredMixin, View):
    """Generates PDF for transferred students report."""

    def get(self, request):
        from .forms import TransferredStudentsFilterForm
        form = TransferredStudentsFilterForm(request.GET)
        form.is_valid()

        cleaned = getattr(form, 'cleaned_data', {})
        academic_year = cleaned.get('academic_year')
        start_date = cleaned.get('start_date')
        end_date = cleaned.get('end_date')
        student_class = cleaned.get('student_class')

        # Base queryset: transferred students
        students_qs = Student.objects.filter(status='transferred').select_related('current_class')

        if academic_year:
            students_qs = students_qs.filter(status_date__year=academic_year.year)
        if start_date:
            students_qs = students_qs.filter(status_date__gte=start_date)
        if end_date:
            students_qs = students_qs.filter(status_date__lte=end_date)
        if student_class:
            students_qs = students_qs.filter(current_class__name=student_class)

        rows = []
        for student in students_qs.order_by('first_name', 'last_name'):
            rows.append({
                'name': student.full_name,
                'admission_number': student.admission_number or 'N/A',
                'grade': student.current_class.name if student.current_class else 'Not assigned',
            })

        context = {
            'rows': rows,
            'filters': {
                'academic_year': academic_year,
                'start_date': start_date,
                'end_date': end_date,
                'student_class': student_class,
            },
            'SCHOOL_NAME': getattr(settings, 'SCHOOL_NAME', 'PCEA Wendani Academy'),
            'SCHOOL_LOGO_URL': request.build_absolute_uri(settings.STATIC_URL + 'assets/images/logo.jpeg'),
            'SPONSOR_LOGO_URL': request.build_absolute_uri(settings.STATIC_URL + 'assets/images/logo2.jpeg'),
            'SCHOOL_ADDRESS': getattr(settings, 'SCHOOL_ADDRESS', ''),
            'SCHOOL_CONTACT': getattr(settings, 'SCHOOL_CONTACT', ''),
            'BANK_DETAILS': getattr(settings, 'SCHOOL_BANK_DETAILS', {
                'equity': {'name': 'EQUITY BANK', 'account_name': 'P.C.E.A Wendani Academy',
                           'account_no': '1130280029105'},
                'coop': {'name': 'CO-OPERATIVE BANK', 'account_name': 'P.C.E.A Wendani Academy',
                         'account_no': '01129158350600'},
                'paybills': [
                    {'label': 'PAYBILL (247247)', 'acc_format': '80029#<admission_number>'},
                    {'label': 'PAYBILL (400222)', 'acc_format': '393939#<admission_number>'},
                ]
            }),
            'generated_by': request.user.get_full_name(),
            'generated_on': datetime.now(),
        }

        html_string = render_to_string('reports/pdf/transferred_students_report_pdf.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
        pdf_bytes = html.write_pdf()
        filename = f"transferred-students-{datetime.now().strftime('%Y%m%d-%H%M')}.pdf"
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ---------- Admitted Students Report Exports ----------
class AdmittedStudentsExcelView(LoginRequiredMixin, View):
    """Exports admitted students report to Excel."""

    def get(self, request):
        from .forms import AdmittedStudentsFilterForm
        from django.utils import timezone as tz
        form = AdmittedStudentsFilterForm(request.GET)
        form.is_valid()

        cleaned = getattr(form, 'cleaned_data', {})
        start_date = cleaned.get('start_date')
        end_date = cleaned.get('end_date')
        student_class = cleaned.get('student_class')

        # Default to current year if no dates provided
        if not start_date:
            start_date = tz.now().replace(month=1, day=1).date()
        if not end_date:
            end_date = tz.now().date()

        # Base queryset: students with admission_date
        students_qs = Student.objects.filter(
            admission_date__isnull=False
        ).select_related('current_class')

        if start_date:
            students_qs = students_qs.filter(admission_date__gte=start_date)
        if end_date:
            students_qs = students_qs.filter(admission_date__lte=end_date)
        if student_class:
            students_qs = students_qs.filter(current_class__name=student_class)

        # Build workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Admitted Students"

        title = "Admitted Students Report"
        if start_date and end_date:
            title += f" ({start_date} to {end_date})"
        add_common_header(ws, title)

        headers = ['Name', 'Admission Number', 'Admission Date', 'Grade']
        for c, h in enumerate(headers, start=1):
            ws.cell(row=5, column=c, value=h).font = Font(bold=True)

        row_num = 6
        for student in students_qs.order_by('admission_date', 'first_name', 'last_name'):
            ws.cell(row=row_num, column=1, value=student.full_name)
            ws.cell(row=row_num, column=2, value=student.admission_number or 'N/A')
            ws.cell(row=row_num, column=3, value=student.admission_date.strftime('%Y-%m-%d') if student.admission_date else '')
            ws.cell(row=row_num, column=4, value=student.current_class.name if student.current_class else 'Not assigned')
            row_num += 1

        # Auto width columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25

        bytes_data = workbook_to_bytes(wb)
        filename = f"admitted-students-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
        return xlsx_response(bytes_data, filename)


class AdmittedStudentsPDFView(LoginRequiredMixin, View):
    """Generates PDF for admitted students report."""

    def get(self, request):
        from .forms import AdmittedStudentsFilterForm
        from django.utils import timezone as tz
        form = AdmittedStudentsFilterForm(request.GET)
        form.is_valid()

        cleaned = getattr(form, 'cleaned_data', {})
        start_date = cleaned.get('start_date')
        end_date = cleaned.get('end_date')
        student_class = cleaned.get('student_class')

        # Default to current year if no dates provided
        if not start_date:
            start_date = tz.now().replace(month=1, day=1).date()
        if not end_date:
            end_date = tz.now().date()

        # Base queryset: students with admission_date
        students_qs = Student.objects.filter(
            admission_date__isnull=False
        ).select_related('current_class')

        if start_date:
            students_qs = students_qs.filter(admission_date__gte=start_date)
        if end_date:
            students_qs = students_qs.filter(admission_date__lte=end_date)
        if student_class:
            students_qs = students_qs.filter(current_class__name=student_class)

        rows = []
        for student in students_qs.order_by('admission_date', 'first_name', 'last_name'):
            rows.append({
                'name': student.full_name,
                'admission_number': student.admission_number or 'N/A',
                'admission_date': student.admission_date,
                'grade': student.current_class.name if student.current_class else 'Not assigned',
            })

        context = {
            'rows': rows,
            'filters': {
                'start_date': start_date,
                'end_date': end_date,
                'student_class': student_class,
            },
            'SCHOOL_NAME': getattr(settings, 'SCHOOL_NAME', 'PCEA Wendani Academy'),
            'SCHOOL_LOGO_URL': request.build_absolute_uri(settings.STATIC_URL + 'assets/images/logo.jpeg'),
            'SPONSOR_LOGO_URL': request.build_absolute_uri(settings.STATIC_URL + 'assets/images/logo2.jpeg'),
            'SCHOOL_ADDRESS': getattr(settings, 'SCHOOL_ADDRESS', ''),
            'SCHOOL_CONTACT': getattr(settings, 'SCHOOL_CONTACT', ''),
            'BANK_DETAILS': getattr(settings, 'SCHOOL_BANK_DETAILS', {
                'equity': {'name': 'EQUITY BANK', 'account_name': 'P.C.E.A Wendani Academy',
                           'account_no': '1130280029105'},
                'coop': {'name': 'CO-OPERATIVE BANK', 'account_name': 'P.C.E.A Wendani Academy',
                         'account_no': '01129158350600'},
                'paybills': [
                    {'label': 'PAYBILL (247247)', 'acc_format': '80029#<admission_number>'},
                    {'label': 'PAYBILL (400222)', 'acc_format': '393939#<admission_number>'},
                ]
            }),
            'generated_by': request.user.get_full_name(),
            'generated_on': datetime.now(),
        }

        html_string = render_to_string('reports/pdf/admitted_students_report_pdf.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
        pdf_bytes = html.write_pdf()
        filename = f"admitted-students-{datetime.now().strftime('%Y%m%d-%H%M')}.pdf"
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ---------- Other Items Report Exports ----------
class OtherItemsReportExcelView(LoginRequiredMixin, OrganizationFilterMixin, View):
    """Exports other items report to Excel."""

    def get(self, request):
        from .forms import OtherItemsReportFilterForm
        form = OtherItemsReportFilterForm(request.GET)
        form.is_valid()

        cleaned = getattr(form, 'cleaned_data', {})
        academic_year = cleaned.get('academic_year')
        term = cleaned.get('term')
        student_class = cleaned.get('student_class') or ''
        name = cleaned.get('name') or ''
        admission = cleaned.get('admission') or ''
        category = cleaned.get('category') or ''
        start_date = cleaned.get('start_date')
        end_date = cleaned.get('end_date')
        show_all = cleaned.get('show_all', False)

        # Base queryset: other items only
        items_qs = InvoiceItem.objects.filter(
            category='other'
        ).select_related('invoice__student', 'invoice__term__academic_year')

        # Apply organization filter
        organization = getattr(request, 'organization', None)
        if organization:
            items_qs = items_qs.filter(invoice__organization=organization)

        # Apply filters
        if not show_all:
            if academic_year:
                items_qs = items_qs.filter(invoice__term__academic_year=academic_year)
                if term:
                    items_qs = items_qs.filter(invoice__term__term=term)
            if student_class:
                items_qs = items_qs.filter(invoice__student__current_class__name=student_class)
            if name:
                items_qs = items_qs.filter(
                    Q(invoice__student__first_name__icontains=name) |
                    Q(invoice__student__middle_name__icontains=name) |
                    Q(invoice__student__last_name__icontains=name)
                )
            if admission:
                items_qs = items_qs.filter(invoice__student__admission_number__icontains=admission)
            if category:
                items_qs = items_qs.filter(description__icontains=category)
            if start_date:
                items_qs = items_qs.filter(invoice__issue_date__gte=start_date)
            if end_date:
                items_qs = items_qs.filter(invoice__issue_date__lte=end_date)

        # Group by student and category (description)
        grouped = items_qs.values(
            'invoice__student__pk',
            'invoice__student__first_name',
            'invoice__student__middle_name',
            'invoice__student__last_name',
            'invoice__student__admission_number',
            'invoice__student__current_class__name',
            'description',
        ).annotate(
            total_billed=Coalesce(Sum('net_amount'), Value(Decimal('0.00')), output_field=DecimalField())
        ).order_by(
            'invoice__student__first_name',
            'invoice__student__last_name',
            'description'
        )

        # Build collected map
        collected_map = {}
        if PaymentAllocation is not None:
            alloc_qs = PaymentAllocation.objects.filter(
                invoice_item__in=items_qs,
                is_active=True,
                payment__is_active=True,
                payment__status='completed'
            ).values(
                'invoice_item__invoice__student__pk',
                'invoice_item__description'
            ).annotate(
                collected=Coalesce(Sum('amount'), Value(Decimal('0.00')), output_field=DecimalField())
            )

            for row in alloc_qs:
                key = (
                    row.get('invoice_item__invoice__student__pk'),
                    row.get('invoice_item__description')
                )
                collected_map[key] = Decimal(row.get('collected') or 0)

        # Build rows
        rows = []
        total_billed = Decimal('0.00')
        total_collected = Decimal('0.00')
        total_balance = Decimal('0.00')

        for g in grouped:
            student_pk = g.get('invoice__student__pk')
            first_name = g.get('invoice__student__first_name') or ''
            middle_name = g.get('invoice__student__middle_name') or ''
            last_name = g.get('invoice__student__last_name') or ''
            student_name = f"{first_name} {middle_name} {last_name}".strip()
            student_name = ' '.join(student_name.split())

            admission = g.get('invoice__student__admission_number') or ''
            student_cls = g.get('invoice__student__current_class__name') or 'Not assigned'
            category_desc = g.get('description') or 'Other'
            billed = Decimal(g.get('total_billed') or 0)
            collected = collected_map.get((student_pk, category_desc), Decimal('0.00'))
            balance = billed - collected

            rows.append({
                'student_name': student_name,
                'admission': admission,
                'student_class': student_cls,
                'category': category_desc,
                'billed': billed,
                'collected': collected,
                'balance': balance
            })

            total_billed += billed
            total_collected += collected
            total_balance += balance

        # Build workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Other Items Report"

        title = "Other Items Invoice Report"
        if academic_year:
            title += f" - {academic_year.year}"
            if term:
                title += f" Term {term}"
        add_common_header(ws, title)

        headers = ['Sr no.', 'Student Name', 'Admission Number', 'Grade', 'Category', 'Total Amount', 'Paid', 'Balance']
        for c, h in enumerate(headers, start=1):
            ws.cell(row=5, column=c, value=h).font = Font(bold=True)

        row_num = 6
        for idx, r in enumerate(rows, start=1):
            ws.cell(row=row_num, column=1, value=idx)
            ws.cell(row=row_num, column=2, value=r['student_name'])
            ws.cell(row=row_num, column=3, value=r['admission'])
            ws.cell(row=row_num, column=4, value=r['student_class'])
            ws.cell(row=row_num, column=5, value=r['category'])

            # Money columns
            c6 = ws.cell(row=row_num, column=6, value=float(r['billed']))
            format_money_cell(c6)
            c7 = ws.cell(row=row_num, column=7, value=float(r['collected']))
            format_money_cell(c7)
            c8 = ws.cell(row=row_num, column=8, value=float(r['balance']))
            format_money_cell(c8)

            row_num += 1

        # Totals
        ws.cell(row=row_num, column=1, value='TOTALS').font = Font(bold=True)
        ws.cell(row=row_num, column=5, value='TOTALS').font = Font(bold=True)
        c6 = ws.cell(row=row_num, column=6, value=float(total_billed))
        format_money_cell(c6)
        c6.font = Font(bold=True)
        c7 = ws.cell(row=row_num, column=7, value=float(total_collected))
        format_money_cell(c7)
        c7.font = Font(bold=True)
        c8 = ws.cell(row=row_num, column=8, value=float(total_balance))
        format_money_cell(c8)
        c8.font = Font(bold=True)

        # Auto width columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        bytes_data = workbook_to_bytes(wb)
        filename = f"other-items-report-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
        return xlsx_response(bytes_data, filename)


class OtherItemsReportPDFView(LoginRequiredMixin, OrganizationFilterMixin, View):
    """Generates PDF for other items report."""

    def get(self, request):
        from .forms import OtherItemsReportFilterForm
        form = OtherItemsReportFilterForm(request.GET)
        form.is_valid()

        cleaned = getattr(form, 'cleaned_data', {})
        academic_year = cleaned.get('academic_year')
        term = cleaned.get('term')
        student_class = cleaned.get('student_class') or ''
        name = cleaned.get('name') or ''
        admission = cleaned.get('admission') or ''
        category = cleaned.get('category') or ''
        start_date = cleaned.get('start_date')
        end_date = cleaned.get('end_date')
        show_all = cleaned.get('show_all', False)

        # Base queryset: other items only
        items_qs = InvoiceItem.objects.filter(
            category='other'
        ).select_related('invoice__student', 'invoice__term__academic_year')

        # Apply organization filter
        organization = getattr(request, 'organization', None)
        if organization:
            items_qs = items_qs.filter(invoice__organization=organization)

        # Apply filters (same logic as Excel view)
        if not show_all:
            if academic_year:
                items_qs = items_qs.filter(invoice__term__academic_year=academic_year)
                if term:
                    items_qs = items_qs.filter(invoice__term__term=term)
            if student_class:
                items_qs = items_qs.filter(invoice__student__current_class__name=student_class)
            if name:
                items_qs = items_qs.filter(
                    Q(invoice__student__first_name__icontains=name) |
                    Q(invoice__student__middle_name__icontains=name) |
                    Q(invoice__student__last_name__icontains=name)
                )
            if admission:
                items_qs = items_qs.filter(invoice__student__admission_number__icontains=admission)
            if category:
                items_qs = items_qs.filter(description__icontains=category)
            if start_date:
                items_qs = items_qs.filter(invoice__issue_date__gte=start_date)
            if end_date:
                items_qs = items_qs.filter(invoice__issue_date__lte=end_date)

        # Group and build rows (same as Excel view)
        grouped = items_qs.values(
            'invoice__student__pk',
            'invoice__student__first_name',
            'invoice__student__middle_name',
            'invoice__student__last_name',
            'invoice__student__admission_number',
            'invoice__student__current_class__name',
            'description',
        ).annotate(
            total_billed=Coalesce(Sum('net_amount'), Value(Decimal('0.00')), output_field=DecimalField())
        ).order_by(
            'invoice__student__first_name',
            'invoice__student__last_name',
            'description'
        )

        collected_map = {}
        if PaymentAllocation is not None:
            alloc_qs = PaymentAllocation.objects.filter(
                invoice_item__in=items_qs,
                is_active=True,
                payment__is_active=True,
                payment__status='completed'
            ).values(
                'invoice_item__invoice__student__pk',
                'invoice_item__description'
            ).annotate(
                collected=Coalesce(Sum('amount'), Value(Decimal('0.00')), output_field=DecimalField())
            )

            for row in alloc_qs:
                key = (
                    row.get('invoice_item__invoice__student__pk'),
                    row.get('invoice_item__description')
                )
                collected_map[key] = Decimal(row.get('collected') or 0)

        rows = []
        total_billed = Decimal('0.00')
        total_collected = Decimal('0.00')
        total_balance = Decimal('0.00')

        for g in grouped:
            student_pk = g.get('invoice__student__pk')
            first_name = g.get('invoice__student__first_name') or ''
            middle_name = g.get('invoice__student__middle_name') or ''
            last_name = g.get('invoice__student__last_name') or ''
            student_name = f"{first_name} {middle_name} {last_name}".strip()
            student_name = ' '.join(student_name.split())

            admission = g.get('invoice__student__admission_number') or ''
            student_cls = g.get('invoice__student__current_class__name') or 'Not assigned'
            category_desc = g.get('description') or 'Other'
            billed = Decimal(g.get('total_billed') or 0)
            collected = collected_map.get((student_pk, category_desc), Decimal('0.00'))
            balance = billed - collected

            rows.append({
                'student_name': student_name,
                'admission': admission,
                'student_class': student_cls,
                'category': category_desc,
                'billed': billed,
                'collected': collected,
                'balance': balance
            })

            total_billed += billed
            total_collected += collected
            total_balance += balance

        context = {
            'rows': rows,
            'totals': {
                'billed': total_billed,
                'collected': total_collected,
                'balance': total_balance
            },
            'filters': {
                'academic_year': academic_year,
                'term': term,
                'student_class': student_class,
                'name': name,
                'admission': admission,
                'category': category,
                'start_date': start_date,
                'end_date': end_date,
            },
            'SCHOOL_NAME': getattr(settings, 'SCHOOL_NAME', 'PCEA Wendani Academy'),
            'SCHOOL_LOGO_URL': getattr(settings, 'SCHOOL_LOGO_URL', '/static/assets/images/logo.jpeg'),
            'SPONSOR_LOGO_URL': getattr(settings, 'SPONSOR_LOGO_URL', '/static/assets/images/logo2.jpeg'),
            'SCHOOL_ADDRESS': getattr(settings, 'SCHOOL_ADDRESS', ''),
            'SCHOOL_CONTACT': getattr(settings, 'SCHOOL_CONTACT', ''),
            'BANK_DETAILS': getattr(settings, 'SCHOOL_BANK_DETAILS', {
                'equity': {'name': 'EQUITY BANK', 'account_name': 'P.C.E.A Wendani Academy',
                           'account_no': '1130280029105'},
                'coop': {'name': 'CO-OPERATIVE BANK', 'account_name': 'P.C.E.A Wendani Academy',
                         'account_no': '01129158350600'},
                'paybills': [
                    {'label': 'PAYBILL (247247)', 'acc_format': '80029#<admission_number>'},
                    {'label': 'PAYBILL (400222)', 'acc_format': '393939#<admission_number>'},
                ]
            }),
            'generated_by': request.user.get_full_name(),
            'generated_on': datetime.now(),
        }

        html_string = render_to_string('reports/pdf/other_items_report_pdf.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
        pdf_bytes = html.write_pdf()
        filename = f"other-items-report-{datetime.now().strftime('%Y%m%d-%H%M')}.pdf"
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

