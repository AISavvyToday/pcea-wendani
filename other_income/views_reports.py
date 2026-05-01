from datetime import datetime
from io import BytesIO

from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.views.generic import TemplateView, View
from openpyxl import Workbook
from openpyxl.styles import Font

from core.mixins import RoleRequiredMixin
from core.models import UserRole
from .forms import OtherIncomeReportStagingFilterForm
from .reporting import (
    build_other_income_report_filters,
    build_other_income_flat_rows,
    build_other_income_summary,
)


class HTML:
    """Lazy WeasyPrint proxy so app checks do not require native PDF libraries."""

    def __init__(self, *args, **kwargs):
        from weasyprint import HTML as WeasyHTML

        self._html = WeasyHTML(*args, **kwargs)

    def write_pdf(self, *args, **kwargs):
        return self._html.write_pdf(*args, **kwargs)


class OtherIncomeReportBaseMixin(LoginRequiredMixin, RoleRequiredMixin):
    allowed_roles = [UserRole.SUPER_ADMIN, UserRole.SCHOOL_ADMIN, UserRole.ACCOUNTANT]

    def get_report_form(self):
        return OtherIncomeReportStagingFilterForm(
            self.request.GET or None,
            organization=getattr(self.request, 'organization', None),
        )

    def get_filters(self):
        form = self.get_report_form()
        form.is_valid()
        return form, build_other_income_report_filters(getattr(form, 'cleaned_data', {}))

    def build_context(self):
        form, filters = self.get_filters()
        organization = getattr(self.request, 'organization', None)
        rows = build_other_income_flat_rows(organization=organization, filters=filters)
        summary = build_other_income_summary(rows)
        return {
            'form': form,
            'rows': rows,
            'summary': summary,
            'organization': organization,
            'generated_at': datetime.now(),
        }


class OtherIncomeReportView(OtherIncomeReportBaseMixin, TemplateView):
    template_name = 'other_income/report_staging.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(self.build_context())
        return context


class OtherIncomeReportExcelView(OtherIncomeReportBaseMixin, View):
    def get(self, request, *args, **kwargs):
        context = self.build_context()
        rows = context['rows']
        summary = context['summary']

        wb = Workbook()
        ws = wb.active
        ws.title = 'Other Income Report'
        headers = [
            'Invoice #', 'Client', 'Contact', 'Description', 'Issue Date', 'Due Date',
            'Status', 'Total Amount', 'Amount Paid', 'Balance', 'Payment Methods',
            'Payments', 'Items', 'Last Payment Date', 'Receipt / Ref'
        ]
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=idx, value=header)
            cell.font = Font(bold=True)

        row_num = 2
        for row in rows:
            ws.cell(row=row_num, column=1, value=row['invoice_number'])
            ws.cell(row=row_num, column=2, value=row['client_name'])
            ws.cell(row=row_num, column=3, value=row['client_contact'])
            ws.cell(row=row_num, column=4, value=row['description'])
            ws.cell(row=row_num, column=5, value=row['issue_date'].strftime('%Y-%m-%d') if row['issue_date'] else '')
            ws.cell(row=row_num, column=6, value=row['due_date'].strftime('%Y-%m-%d') if row['due_date'] else '')
            ws.cell(row=row_num, column=7, value=row['status'].replace('_', ' ').title())
            ws.cell(row=row_num, column=8, value=float(row['total_amount']))
            ws.cell(row=row_num, column=9, value=float(row['amount_paid']))
            ws.cell(row=row_num, column=10, value=float(row['balance']))
            ws.cell(row=row_num, column=11, value=row['payment_methods_display'])
            ws.cell(row=row_num, column=12, value=row['payment_count'])
            ws.cell(row=row_num, column=13, value=row['item_count'])
            ws.cell(row=row_num, column=14, value=row['last_payment_date'].strftime('%Y-%m-%d %H:%M') if row['last_payment_date'] else '')
            ws.cell(row=row_num, column=15, value=row['last_reference'])
            row_num += 1

        row_num += 1
        ws.cell(row=row_num, column=1, value='Summary').font = Font(bold=True)
        ws.cell(row=row_num + 1, column=1, value='Invoice Count')
        ws.cell(row=row_num + 1, column=2, value=summary['invoice_count'])
        ws.cell(row=row_num + 2, column=1, value='Total Invoiced')
        ws.cell(row=row_num + 2, column=2, value=float(summary['total_invoiced']))
        ws.cell(row=row_num + 3, column=1, value='Total Paid')
        ws.cell(row=row_num + 3, column=2, value=float(summary['total_paid']))
        ws.cell(row=row_num + 4, column=1, value='Total Balance')
        ws.cell(row=row_num + 4, column=2, value=float(summary['total_balance']))

        for column_cells in ws.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 30)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="other-income-report-{datetime.now().strftime("%Y%m%d-%H%M")}.xlsx"'
        wb.save(response)
        return response


class OtherIncomeReportPDFView(OtherIncomeReportBaseMixin, View):
    def get(self, request, *args, **kwargs):
        context = self.build_context()
        html_string = render_to_string('other_income/pdf/report_pdf.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
        pdf_bytes = html.write_pdf()
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="other-income-report-{datetime.now().strftime("%Y%m%d-%H%M")}.pdf"'
        return response
