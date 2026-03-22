# students/views_exports.py
"""
Export views for student list - Excel and PDF downloads.
"""
import io
from datetime import datetime
from decimal import Decimal

from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.views import View

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from weasyprint import HTML

from core.mixins import RoleRequiredMixin
from core.models import UserRole
from .models import Student
from .services import StudentService


def workbook_to_bytes(wb):
    """Convert openpyxl workbook to bytes."""
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def xlsx_response(workbook_bytes, filename):
    """Return HTTP response with Excel file."""
    response = HttpResponse(
        workbook_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


class StudentListExcelView(LoginRequiredMixin, RoleRequiredMixin, View):
    """Export student list to Excel with applied filters."""
    
    allowed_roles = [
        UserRole.SUPER_ADMIN,
        UserRole.SCHOOL_ADMIN,
        UserRole.ACCOUNTANT,
        UserRole.TEACHER
    ]

    def get(self, request):
        # Apply the same filters as the list view
        query = request.GET.get('query', '')
        class_id = request.GET.get('current_class', '')
        # Get status parameter - default to 'active' if not provided (matching list view behavior)
        status = request.GET.get('status', '') or 'active'
        gender = request.GET.get('gender', '')
        is_boarder = request.GET.get('is_boarder', '')
        stream = request.GET.get('stream', '')
        club_id = request.GET.get('club', '')
        organization = getattr(request, 'organization', None)

        queryset = StudentService.search_students(
            query=query if query else None,
            class_id=class_id if class_id else None,
            status=status,  # Use the status from URL parameters
            gender=gender if gender else None,
            is_boarder=is_boarder if is_boarder else None,
            stream=stream if stream else None,
            club_id=club_id if club_id else None,
            organization=organization
        )

        # Build workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"

        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # School header
        ws.merge_cells('A1:J1')
        ws['A1'] = getattr(settings, 'SCHOOL_NAME', 'PCEA Wendani Academy')
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:J2')
        ws['A2'] = f"Student List Report - Generated on {datetime.now().strftime('%d %b %Y %H:%M')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')

        # Filters applied info
        ws.merge_cells('A3:J3')
        filter_info = []
        if query:
            filter_info.append(f"Search: {query}")
        if class_id:
            from academics.models import Class
            try:
                cls = Class.objects.get(pk=class_id)
                filter_info.append(f"Class: {cls}")
            except Class.DoesNotExist:
                filter_info.append(f"Class ID: {class_id}")
        if status:
            filter_info.append(f"Status: {status}")
        if gender:
            filter_info.append(f"Gender: {gender}")
        if is_boarder:
            filter_info.append(f"Boarding: {is_boarder}")
        if stream:
            filter_info.append(f"Stream: {stream}")
        if club_id:
            filter_info.append("Club filter applied")
        ws['A3'] = f"Filters: {', '.join(filter_info) if filter_info else 'All Students'}"
        ws['A3'].font = Font(size=9)
        ws['A3'].alignment = Alignment(horizontal='center')

        # Column headers
        headers = [
            'Admission No.',
            'Full Name',
            'Gender',
            'Date of Birth',
            'Class',
            'Status',
            'Transport Route',
            'Primary Parent',
            'Parent Phone',
            'Admission Date'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        row_num = 6
        for student in queryset:
            primary_parent = student.primary_parent
            
            row_data = [
                student.admission_number,
                student.full_name,
                student.get_gender_display() if hasattr(student, 'get_gender_display') else student.gender,
                student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '',
                str(student.current_class) if student.current_class else 'Not Assigned',
                student.get_status_display() if hasattr(student, 'get_status_display') else student.status,
                str(student.transport_route) if student.transport_route else '',
                primary_parent.full_name if primary_parent else '',
                primary_parent.phone_primary if primary_parent else '',
                student.admission_date.strftime('%Y-%m-%d') if student.admission_date else '',
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')

            row_num += 1

        # Summary row
        ws.cell(row=row_num + 1, column=1, value=f"Total Students: {queryset.count()}")
        ws.cell(row=row_num + 1, column=1).font = Font(bold=True)

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(5, row_num + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Generate response with filter info in filename
        bytes_data = workbook_to_bytes(wb)
        filename_parts = ["students-report"]
        if status and status != 'active':
            filename_parts.append(status)
        if class_id:
            filename_parts.append(f"class-{class_id}")
        filename_parts.append(datetime.now().strftime('%Y%m%d-%H%M'))
        filename = f"{'-'.join(filename_parts)}.xlsx"
        return xlsx_response(bytes_data, filename)


class StudentListPDFView(LoginRequiredMixin, RoleRequiredMixin, View):
    """Export student list to PDF with applied filters."""
    
    allowed_roles = [
        UserRole.SUPER_ADMIN,
        UserRole.SCHOOL_ADMIN,
        UserRole.ACCOUNTANT,
        UserRole.TEACHER
    ]

    def get(self, request):
        # Apply the same filters as the list view
        query = request.GET.get('query', '')
        class_id = request.GET.get('current_class', '')
        # Get status parameter - default to 'active' if not provided (matching list view behavior)
        status = request.GET.get('status', '') or 'active'
        gender = request.GET.get('gender', '')
        is_boarder = request.GET.get('is_boarder', '')
        stream = request.GET.get('stream', '')
        club_id = request.GET.get('club', '')
        organization = getattr(request, 'organization', None)

        queryset = StudentService.search_students(
            query=query if query else None,
            class_id=class_id if class_id else None,
            status=status,  # Use the status from URL parameters
            gender=gender if gender else None,
            is_boarder=is_boarder if is_boarder else None,
            stream=stream if stream else None,
            club_id=club_id if club_id else None,
            organization=organization
        )

        # Prepare filter info
        filter_info = []
        if query:
            filter_info.append(f"Search: {query}")
        if class_id:
            from academics.models import Class
            try:
                cls = Class.objects.get(pk=class_id)
                filter_info.append(f"Class: {cls}")
            except Class.DoesNotExist:
                filter_info.append(f"Class: {class_id}")
        if status:
            filter_info.append(f"Status: {status}")
        if gender:
            filter_info.append(f"Gender: {gender}")
        if is_boarder:
            filter_info.append(f"Boarding: {is_boarder}")
        if stream:
            filter_info.append(f"Stream: {stream}")
        if club_id:
            filter_info.append("Club filter applied")

        # Build context
        context = {
            'students': queryset,
            'total_students': queryset.count(),
            'filters_applied': ', '.join(filter_info) if filter_info else 'None',
            'SCHOOL_NAME': getattr(settings, 'SCHOOL_NAME', 'PCEA Wendani Academy'),
            'SCHOOL_LOGO_URL': request.build_absolute_uri(settings.STATIC_URL + 'assets/images/logo.jpeg'),
            'SPONSOR_LOGO_URL': request.build_absolute_uri(settings.STATIC_URL + 'assets/images/logo2.jpeg'),
            'SCHOOL_ADDRESS': getattr(settings, 'SCHOOL_ADDRESS', ''),
            'SCHOOL_CONTACT': getattr(settings, 'SCHOOL_CONTACT', ''),
            'generated_by': request.user.get_full_name() if hasattr(request.user, 'get_full_name') else str(request.user),
            'generated_on': datetime.now(),
        }

        html_string = render_to_string('students/pdf/student_list_pdf.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
        pdf_bytes = html.write_pdf()

        # Generate filename with filter info
        filename_parts = ["students-report"]
        if status and status != 'active':
            filename_parts.append(status)
        if class_id:
            filename_parts.append(f"class-{class_id}")
        filename_parts.append(datetime.now().strftime('%Y%m%d-%H%M'))
        filename = f"{'-'.join(filename_parts)}.pdf"
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


