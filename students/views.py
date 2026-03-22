# students/views.py
import logging
from decimal import Decimal

from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.db.models import Q, Sum
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse, Http404
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import ListView, CreateView, UpdateView, DetailView, FormView, View
from django.core.paginator import Paginator
from django.db import transaction
from django.views.generic import DeleteView
from django.contrib import messages
from core.mixins import RoleRequiredMixin, OrganizationFilterMixin
from accounts.models import User
from .models import Student, Parent, StudentParent
from .forms import StudentForm, ParentForm, StudentSearchForm, StudentPromotionForm, StudentImportForm
from .metrics import apply_student_filters, get_current_term, get_student_base_queryset, get_student_status_counters
from .services import StudentService
from academics.models import Class, AcademicYear, Term
from core.models import UserRole, InvoiceStatus
from finance.models import Invoice

logger = logging.getLogger(__name__)


class ParentOrganizationQuerysetMixin:
    """Centralize organization-safe parent lookups for parent management views."""

    def get_parent_queryset(self):
        queryset = Parent.objects.all()
        organization = getattr(self.request, 'organization', None)

        if organization is None:
            return queryset.none()

        return queryset.filter(organization=organization)


class StudentListView(LoginRequiredMixin, OrganizationFilterMixin, RoleRequiredMixin, ListView):
    """List all students with search and filters"""

    model = Student
    template_name = 'students/student_list.html'
    context_object_name = 'students'
    paginate_by = 20
    allowed_roles = [
        UserRole.SUPER_ADMIN,
        UserRole.SCHOOL_ADMIN,
        UserRole.ACCOUNTANT,
        UserRole.TEACHER
    ]

    def get_queryset(self):
        base_queryset = self.get_base_filtered_queryset()
        status = self.request.GET.get('status', 'active') or 'active'
        term = self.get_current_term()
        queryset = apply_student_filters(
            base_queryset,
            status=status,
            term=term,
        )
        return queryset.order_by('admission_number')

    def get_current_term(self):
        if not hasattr(self, '_current_term'):
            organization = getattr(self.request, 'organization', None)
            self._current_term = get_current_term(organization=organization)
        return self._current_term

    def get_base_filtered_queryset(self):
        organization = getattr(self.request, 'organization', None)
        queryset = get_student_base_queryset(organization=organization)
        queryset = queryset.select_related('current_class').prefetch_related('parents')
        return apply_student_filters(
            queryset,
            query=self.request.GET.get('query', '') or None,
            class_id=self.request.GET.get('current_class', '') or None,
            status=None,
            gender=self.request.GET.get('gender', '') or None,
            is_boarder=self.request.GET.get('is_boarder', '') or None,
            stream=self.request.GET.get('stream', '') or None,
        )

    def get_paginate_by(self, queryset):
        per_page = self.request.GET.get('per_page')
        try:
            per_page = int(per_page)
            if per_page <= 0:
                raise ValueError()
        except Exception:
            per_page = self.paginate_by
        return per_page

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        base_queryset = self.get_base_filtered_queryset()
        term = self.get_current_term()
        status_counts = get_student_status_counters(base_queryset, term=term)

        context['search_form'] = StudentSearchForm(self.request.GET)
        context['total_students'] = base_queryset.count()
        context['status_counts'] = status_counts
        context['active_students'] = status_counts['active']

        # Get current status from request (default to 'active' if not specified)
        context['current_status'] = self.request.GET.get('status', '') or 'active'

        return context


class StudentCreateView(LoginRequiredMixin, OrganizationFilterMixin, RoleRequiredMixin, CreateView):
    """Create a new student"""

    model = Student
    form_class = StudentForm
    template_name = 'students/student_form.html'
    success_url = reverse_lazy('students:list')
    allowed_roles = [UserRole.SUPER_ADMIN, UserRole.SCHOOL_ADMIN]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Register New Student'
        context['button_text'] = 'Register Student'
        context['is_edit'] = False  # Always False for create view

        if self.request.POST:
            # Check if parent forms have any data before binding
            parent_1_has_data = bool(
                self.request.POST.get('parent1-first_name', '').strip() or 
                self.request.POST.get('parent1-last_name', '').strip() or
                self.request.POST.get('parent1-phone_primary', '').strip()
            )
            parent_2_has_data = bool(
                self.request.POST.get('parent2-first_name', '').strip() or 
                self.request.POST.get('parent2-last_name', '').strip() or
                self.request.POST.get('parent2-phone_primary', '').strip()
            )
            
            # Only bind forms if they have data, otherwise use unbound forms
            if parent_1_has_data:
                context['parent_form_1'] = ParentForm(self.request.POST, prefix='parent1')
            else:
                context['parent_form_1'] = ParentForm(prefix='parent1')
            
            if parent_2_has_data:
                context['parent_form_2'] = ParentForm(self.request.POST, prefix='parent2')
            else:
                context['parent_form_2'] = ParentForm(prefix='parent2')
        else:
            context['parent_form_1'] = ParentForm(prefix='parent1')
            context['parent_form_2'] = ParentForm(prefix='parent2')

        return context

    def form_valid(self, form):
        context = self.get_context_data()
        parent_form_1 = context['parent_form_1']
        parent_form_2 = context['parent_form_2']

        # Check if parent forms have any data by checking POST data directly
        parent_1_has_data = bool(
            self.request.POST.get('parent1-first_name', '').strip() or 
            self.request.POST.get('parent1-last_name', '').strip() or
            self.request.POST.get('parent1-phone_primary', '').strip()
        )
        
        parent_2_has_data = bool(
            self.request.POST.get('parent2-first_name', '').strip() or 
            self.request.POST.get('parent2-last_name', '').strip() or
            self.request.POST.get('parent2-phone_primary', '').strip()
        )
        
        # Validate parent forms only if they have data
        # If no data is provided, skip validation entirely (parents are optional)
        parent_1_valid = True
        parent_2_valid = True
        
        if parent_1_has_data:
            parent_1_valid = parent_form_1.is_valid()
            if not parent_1_valid:
                for field, errors in parent_form_1.errors.items():
                    for error in errors:
                        messages.error(self.request, f'Parent 1 - {field}: {error}')
        else:
            # No data for parent 1 - mark as valid and skip validation
            parent_1_valid = True
        
        if parent_2_has_data:
            parent_2_valid = parent_form_2.is_valid()
            if not parent_2_valid:
                for field, errors in parent_form_2.errors.items():
                    for error in errors:
                        messages.error(self.request, f'Parent 2 - {field}: {error}')
        else:
            # No data for parent 2 - mark as valid and skip validation
            parent_2_valid = True
        
        # Parent information is now optional - no need to check if at least one is provided
        
        # If validation failed, return invalid
        if not parent_1_valid or not parent_2_valid:
            return self.form_invalid(form)

        # Save student form (admission_number will be auto-generated in model's save() if not set)
        # Use commit=False to get instance without saving, then add parents
        student = form.save(commit=False)
        
        # Ensure status is set to 'active' for new students
        if not student.status:
            student.status = 'active'
        
        # Save student first (admission_number will be auto-generated in model's save() if not provided)
        student.save()

        # Prepare parents data
        parents_data = []

        if parent_1_has_data and parent_1_valid:
            parent_1_data = parent_form_1.cleaned_data.copy()
            parent_1_data['is_primary'] = True
            parent_1_data['relationship'] = parent_1_data.get('relationship', 'guardian')
            parents_data.append(parent_1_data)

        if parent_2_has_data and parent_2_valid:
            parent_2_data = parent_form_2.cleaned_data.copy()
            parent_2_data['is_primary'] = False
            parent_2_data['relationship'] = parent_2_data.get('relationship', 'guardian')
            parents_data.append(parent_2_data)

        try:
            # Add parents to the student using service
            if parents_data:
                for parent_data in parents_data:
                    # Extract parent info and relationship data
                    parent_info = {}
                    relationship_data = {}
                    
                    # Parent fields
                    parent_fields = ['first_name', 'last_name', 'gender', 'id_number', 
                                   'phone_primary', 'phone_secondary', 'email', 'address', 
                                   'town', 'occupation', 'employer']
                    
                    for field in parent_fields:
                        if field in parent_data:
                            parent_info[field] = parent_data.pop(field)
                    
                    # Relationship fields
                    relationship_fields = ['relationship', 'is_primary', 'is_emergency_contact', 
                                         'can_pickup', 'receives_notifications']
                    for field in relationship_fields:
                        if field in parent_data:
                            relationship_data[field] = parent_data.pop(field)

                    # Check if parent already exists by phone or ID
                    parent = None
                    if parent_info.get('phone_primary'):
                        parent = Parent.objects.filter(
                            phone_primary=parent_info['phone_primary']
                        ).first()

                    if not parent and parent_info.get('id_number'):
                        parent = Parent.objects.filter(
                            id_number=parent_info['id_number']
                        ).first()

                    # Create parent if doesn't exist
                    if not parent:
                        parent = Parent.objects.create(**parent_info)

                    # Create StudentParent relationship
                    StudentParent.objects.create(
                        student=student,
                        parent=parent,
                        relationship=relationship_data.get('relationship', 'guardian'),
                        is_primary=relationship_data.get('is_primary', False),
                        is_emergency_contact=relationship_data.get('is_emergency_contact', False),
                        can_pickup=relationship_data.get('can_pickup', True),
                        receives_notifications=relationship_data.get('receives_notifications', True),
                    )

            messages.success(
                self.request,
                f'Student {student.full_name} registered successfully with admission number {student.admission_number}!'
            )
            return redirect(self.success_url)

        except Exception as e:
            import traceback
            error_msg = str(e)
            logger.error(f"Error registering student: {error_msg}\n{traceback.format_exc()}")
            messages.error(self.request, f'Error registering student: {error_msg}')
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        """Handle invalid form submission with proper context."""
        context = self.get_context_data()
        context['form'] = form
        return self.render_to_response(context)


class StudentUpdateView(LoginRequiredMixin, OrganizationFilterMixin, RoleRequiredMixin, UpdateView):
    """Edit existing student - preserves financial records on status change"""

    model = Student
    form_class = StudentForm
    template_name = 'students/student_form.html'
    allowed_roles = [UserRole.SUPER_ADMIN, UserRole.SCHOOL_ADMIN]

    def get_success_url(self):
        return reverse_lazy('students:detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Student: {self.object.full_name}'
        context['button_text'] = 'Update Student'
        context['is_edit'] = True
        context['original_status'] = self.object.status
        return context

    def form_valid(self, form):
        # Track status changes
        original_status = self.get_object().status
        new_status = form.cleaned_data.get('status')
        
        if original_status != new_status:
            # Status is changing - update status_date
            form.instance.status_date = timezone.now()
            
            # Handle invoice deletion and balance_bf restoration for graduated/transferred students
            if new_status in ['graduated', 'transferred']:
                student = form.instance
                
                # Get ALL active invoices across ALL terms (not just current term)
                # Transferred/graduated students should have NO active invoices
                all_active_invoices = Invoice.objects.filter(
                    student=student,
                    is_active=True
                ).exclude(status=InvoiceStatus.CANCELLED)
                
                if all_active_invoices.exists():
                    # Restore frozen fields from invoices before deactivating them
                    current_credit = student.credit_balance or Decimal('0.00')
                    total_balance_bf = Decimal('0.00')
                    total_prepayment = Decimal('0.00')
                    
                    for invoice in all_active_invoices:
                        # Restore balance_bf_original to Student frozen field
                        # Note: balance_bf is debt, not credit, so it's NOT added to credit_balance
                        if invoice.balance_bf_original and invoice.balance_bf_original > 0:
                            if form.instance.balance_bf_original == Decimal('0.00'):
                                form.instance.balance_bf_original = invoice.balance_bf_original
                            else:
                                form.instance.balance_bf_original += invoice.balance_bf_original
                            total_balance_bf += invoice.balance_bf_original
                        elif invoice.balance_bf and invoice.balance_bf > 0:
                            # Fallback if balance_bf_original not set
                            if form.instance.balance_bf_original == Decimal('0.00'):
                                form.instance.balance_bf_original = invoice.balance_bf
                            else:
                                form.instance.balance_bf_original += invoice.balance_bf
                            total_balance_bf += invoice.balance_bf
                        
                        # Restore prepayment_original to Student frozen field
                        # Prepayment is stored as POSITIVE credit
                        if invoice.prepayment and invoice.prepayment > 0:
                            prepayment_amount = invoice.prepayment
                            if form.instance.prepayment_original == Decimal('0.00'):
                                form.instance.prepayment_original = prepayment_amount
                            else:
                                form.instance.prepayment_original += prepayment_amount
                            current_credit += prepayment_amount
                            total_prepayment += prepayment_amount
                    
                    # Soft delete ALL active invoices (set is_active=False)
                    # This ensures transferred/graduated students have NO active invoices
                    all_active_invoices.update(is_active=False)
                    
                    # Update student's credit_balance and frozen fields
                    form.instance.credit_balance = current_credit
                    
                    # Build message
                    msg_parts = [f'Student status changed to "{new_status}". All active invoices deactivated.']
                    if total_balance_bf > 0:
                        msg_parts.append(f'Balance b/f ({total_balance_bf:,.2f}) restored.')
                    if total_prepayment > 0:
                        msg_parts.append(f'Prepayment ({total_prepayment:,.2f}) restored.')
                    
                    messages.info(self.request, ' '.join(msg_parts))
            
            # Handle invoice restoration when student is reactivated
            # This prevents the bug where invoices remain inactive after status is changed back
            if original_status in ['graduated', 'transferred', 'suspended', 'expelled', 'withdrawn', 'inactive'] and new_status == 'active':
                student = form.instance
                current_term = Term.objects.filter(is_current=True).first()
                
                if current_term:
                    # Restore soft-deleted invoices for current term
                    inactive_invoices = Invoice.objects.filter(
                        student=student,
                        term=current_term,
                        is_active=False
                    ).exclude(status=InvoiceStatus.CANCELLED)
                    
                    restored_count = inactive_invoices.update(is_active=True)
                    
                    if restored_count > 0:
                        # Recompute student balances after restoring invoices
                        student.recompute_outstanding_balance()
                        messages.info(
                            self.request,
                            f'{restored_count} invoice(s) restored for current term. '
                            f'Outstanding balance recalculated.'
                        )
            
            # Add a note about who changed the status
            current_reason = form.instance.status_reason or ''
            change_note = f"Status changed from {original_status} to {new_status} by {self.request.user.get_full_name()} on {timezone.now().strftime('%Y-%m-%d %H:%M')}"
            if current_reason:
                form.instance.status_reason = f"{change_note}\n---\n{current_reason}"
            else:
                form.instance.status_reason = change_note
            
            if new_status not in ['graduated', 'transferred']:
                messages.info(
                    self.request,
                    f'Student status changed from "{original_status}" to "{new_status}". '
                    f'Financial records have been preserved.'
                )
        
        messages.success(self.request, 'Student updated successfully!')
        return super().form_valid(form)



class StudentDetailView(LoginRequiredMixin, OrganizationFilterMixin, RoleRequiredMixin, DetailView):
    model = Student
    template_name = 'students/student_detail.html'
    context_object_name = 'student'
    allowed_roles = [
        UserRole.SUPER_ADMIN,
        UserRole.SCHOOL_ADMIN,
        UserRole.ACCOUNTANT,
        UserRole.TEACHER
    ]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.object

        # ----------------------------
        # CORE PROFILE DATA (SERVICE)
        # ----------------------------
        profile_data = StudentService.get_student_profile_data(student)
        context.update(profile_data)

        # ----------------------------
        # FINANCE DATA
        # ----------------------------
        invoices = student.invoices.filter(is_active=True).select_related('term').order_by('-issue_date')[:25]
        # Add helper fields to each invoice for template display (keep in sync with finance.InvoiceListView)
        for inv in invoices:
            inv.prepayment_abs = abs(inv.prepayment) if inv.prepayment else Decimal('0.00')
            inv.total_due = (
                (inv.total_amount or Decimal("0.00"))
                + (inv.balance_bf or Decimal("0.00"))
                - (inv.prepayment or Decimal("0.00"))
            )
        
        payments = student.payments.order_by('-payment_date')[:25]

        context['invoices'] = invoices
        context['payments'] = payments

        # Total paid (completed payments only)
        total_paid = student.payments.filter(
            status='completed'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        context['credit_balance'] = abs(student.credit_balance)
        context['total_paid'] = total_paid

        

        context['outstanding_balance'] = student.outstanding_balance


        # ----------------------------
        # DOCUMENTS & RECORDS
        # ----------------------------
        context['documents'] = student.documents.order_by('-created_at')[:50]
        context['discipline_records'] = student.discipline_records.order_by('-incident_date')[:50]
        context['medical_records'] = student.medical_records.order_by('-record_date')[:50]

        context['grades'] = student.grades.select_related(
            'exam', 'subject'
        ).order_by('-entered_at')[:100]

        context['attendance_recent'] = student.attendance_records.select_related(
            'class_obj'
        ).order_by('-date')[:50]

        # Enrollment history (from service)
        context['enrollments'] = profile_data.get('enrollments', [])

        # Parents (from service)
        context['student_parents'] = profile_data.get('student_parents', [])

        # Active tab (server-side fallback)
        context['active_tab'] = self.request.GET.get('tab', 'overview')

        return context

class StudentPromotionView(LoginRequiredMixin, OrganizationFilterMixin, RoleRequiredMixin, FormView):
    """Bulk promote students to next class"""

    template_name = 'students/student_promotion.html'
    form_class = StudentPromotionForm
    success_url = reverse_lazy('students:list')
    allowed_roles = [UserRole.SUPER_ADMIN, UserRole.SCHOOL_ADMIN]

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()

        # Get students from selected class
        class_id = self.request.GET.get('class_id')
        if class_id:
            students = Student.objects.filter(
                current_class_id=class_id,
                status='active'
            )
        else:
            students = Student.objects.filter(status='active')

        kwargs['students'] = students
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['classes'] = Class.objects.all()
        context['selected_class'] = self.request.GET.get('class_id')
        context['current_year'] = AcademicYear.objects.filter(is_current=True).first()
        context['current_term'] = Term.objects.filter(is_current=True).first()
        return context

    def form_valid(self, form):
        student_ids = form.cleaned_data['student_ids']
        target_class = form.cleaned_data['target_class']

        # Get current academic year and term
        academic_year = AcademicYear.objects.filter(is_current=True).first()
        term = Term.objects.filter(is_current=True).first()

        if not academic_year or not term:
            messages.error(self.request, 'No current academic year or term set.')
            return redirect('students:promote')

        try:
            promoted_count = StudentService.promote_students(
                student_ids=student_ids,
                target_class=target_class,
                academic_year=academic_year,
                term=term
            )

            messages.success(
                self.request,
                f'Successfully promoted {promoted_count} student(s) to {target_class.name}!'
            )
        except Exception as e:
            messages.error(self.request, f'Error promoting students: {str(e)}')

        return redirect(self.success_url)


class StudentDeleteView(LoginRequiredMixin, OrganizationFilterMixin, RoleRequiredMixin, DeleteView):
    """
    View for soft-deleting a student record.
    Only accessible by Super Admin and School Admin.
    """
    model = Student
    template_name = 'students/student_confirm_delete.html'
    success_url = reverse_lazy('students:list')
    context_object_name = 'student'

    allowed_roles = [
        UserRole.SUPER_ADMIN,
        UserRole.SCHOOL_ADMIN,
    ]

    def form_valid(self, form):
        """
        Perform soft delete instead of hard delete.
        Changes student status to 'inactive' instead of removing from database.
        """
        self.object = self.get_object()
        success_url = self.get_success_url()

        # Soft delete - change status to inactive
        self.object.status = 'inactive'
        self.object.status_date = timezone.now()
        self.object.status_reason = f"Deleted by {self.request.user.get_full_name()}"
        self.object.save()

        messages.success(
            self.request,
            f'Student {self.object.full_name} ({self.object.admission_number}) has been deactivated successfully.'
        )

        return HttpResponseRedirect(success_url)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Delete Student'

        # Get related data to show what will be affected
        student = self.object
        context['related_data'] = {
            'parents': student.parents.count(),
            'invoices': student.invoices.count(),
            'payments': student.payments.count(),
            'attendance_records': student.attendance_records.count(),
        }

        return context


# students/views.py - Add these views at the end

class ParentListView(LoginRequiredMixin, ParentOrganizationQuerysetMixin, OrganizationFilterMixin, RoleRequiredMixin, ListView):
    """List all parents/guardians with search"""

    model = Parent
    template_name = 'students/parent_list.html'
    context_object_name = 'parents'
    paginate_by = 20
    allowed_roles = [
        UserRole.SUPER_ADMIN,
        UserRole.SCHOOL_ADMIN,
        UserRole.ACCOUNTANT,
        UserRole.TEACHER
    ]

    def get_queryset(self):
        queryset = super().get_queryset().prefetch_related('children')
        queryset = self.get_parent_queryset().prefetch_related('children')

        # Search functionality
        query = self.request.GET.get('query', '').strip()
        if query:
            queryset = queryset.filter(
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(phone_primary__icontains=query) |
                Q(id_number__icontains=query) |
                Q(email__icontains=query)
            )

        # Filter by relationship
        relationship = self.request.GET.get('relationship', '').strip()
        if relationship:
            queryset = queryset.filter(relationship=relationship)

        return queryset.order_by('last_name', 'first_name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_parents'] = self.get_queryset().count()
        parent_queryset = self.get_parent_queryset()
        context['total_parents'] = parent_queryset.count()
        context['query'] = self.request.GET.get('query', '')
        context['relationship_filter'] = self.request.GET.get('relationship', '')
        context['relationship_choices'] = Parent.RELATIONSHIP_CHOICES
        return context


class ParentDetailView(LoginRequiredMixin, ParentOrganizationQuerysetMixin, OrganizationFilterMixin, RoleRequiredMixin, DetailView):
    """Parent profile with their children"""

    model = Parent
    template_name = 'students/parent_detail.html'
    context_object_name = 'parent'
    allowed_roles = [
        UserRole.SUPER_ADMIN,
        UserRole.SCHOOL_ADMIN,
        UserRole.ACCOUNTANT,
        UserRole.TEACHER
    ]

    def get_queryset(self):
        return self.get_parent_queryset().prefetch_related(
            'children',
            'parent_students__student',
            'parent_students__student__current_class',
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get all children with their relationships
        context['children'] = StudentParent.objects.filter(
            parent=self.object,
            student__organization=getattr(self.request, 'organization', None),
        ).select_related('student', 'student__current_class')

        # Get total outstanding balance for all children
        total_balance = 0
        for sp in context['children']:
            # This will be implemented when finance module is ready
            # total_balance += sp.student.get_outstanding_balance()
            pass

        context['total_balance'] = total_balance
        context['active_tab'] = self.request.GET.get('tab', 'overview')

        return context


class ParentCreateView(LoginRequiredMixin, OrganizationFilterMixin, RoleRequiredMixin, CreateView):
    """Create a new parent/guardian"""

    model = Parent
    form_class = ParentForm
    template_name = 'students/parent_form.html'
    success_url = reverse_lazy('students:parent_list')
    allowed_roles = [UserRole.SUPER_ADMIN, UserRole.SCHOOL_ADMIN]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Register New Parent/Guardian'
        context['button_text'] = 'Register Parent'
        return context

    def form_valid(self, form):
        messages.success(
            self.request,
            f'Parent/Guardian {form.instance.full_name} registered successfully!'
        )
        return super().form_valid(form)


class ParentUpdateView(LoginRequiredMixin, ParentOrganizationQuerysetMixin, OrganizationFilterMixin, RoleRequiredMixin, UpdateView):
    """Edit existing parent/guardian"""

    model = Parent
    form_class = ParentForm
    template_name = 'students/parent_form.html'
    allowed_roles = [UserRole.SUPER_ADMIN, UserRole.SCHOOL_ADMIN]

    def get_queryset(self):
        return self.get_parent_queryset()

    def get_success_url(self):
        return reverse_lazy('students:parent_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Parent: {self.object.full_name}'
        context['button_text'] = 'Update Parent'
        context['is_edit'] = True
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Parent/Guardian updated successfully!')
        return super().form_valid(form)


class ParentDeleteView(LoginRequiredMixin, ParentOrganizationQuerysetMixin, OrganizationFilterMixin, RoleRequiredMixin, DeleteView):
    """Soft delete a parent/guardian"""

    model = Parent
    template_name = 'students/parent_confirm_delete.html'
    success_url = reverse_lazy('students:parent_list')
    context_object_name = 'parent'
    allowed_roles = [UserRole.SUPER_ADMIN, UserRole.SCHOOL_ADMIN]

    def get_queryset(self):
        return self.get_parent_queryset().prefetch_related('children')

    def form_valid(self, form):
        self.object = self.get_object()
        success_url = self.get_success_url()

        # Check if parent has children
        children_count = self.object.children.count()
        if children_count > 0:
            messages.error(
                self.request,
                f'Cannot delete {self.object.full_name}. They have {children_count} child(ren) linked. Please unlink children first.'
            )
            return redirect('students:parent_detail', pk=self.object.pk)

        # Delete parent
        parent_name = self.object.full_name
        self.object.delete()

        messages.success(
            self.request,
            f'Parent/Guardian {parent_name} has been deleted successfully.'
        )

        return HttpResponseRedirect(success_url)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Delete Parent/Guardian'
        context['children'] = self.object.children.all()
        return context


class ParentChildrenAPIView(LoginRequiredMixin, View):
    """API endpoint to get parent's children with their outstanding balances."""

    def get(self, request, pk):
        from django.http import JsonResponse
        from django.db.models import Sum
        from finance.models import Invoice

        organization = getattr(request, 'organization', None)
        if organization is None:
            raise Http404('Organization not found for request')

        try:
            parent = Parent.objects.get(pk=pk, organization=organization)
        except Parent.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Parent not found'}, status=404)

        # Get all children linked to this parent
        student_parents = StudentParent.objects.filter(
            parent=parent,
            student__organization=organization,
        ).select_related(
            'student', 'student__current_class'
        )

        children_data = []
        for sp in student_parents:
            student = sp.student
            if student.status != 'active':
                continue

            # Calculate outstanding balance
            outstanding = Invoice.objects.filter(
                student=student,
                is_active=True,
                balance__gt=0
            ).aggregate(total=Sum('balance'))['total'] or 0

            children_data.append({
                'id': str(student.pk),
                'name': student.full_name,
                'admission_number': student.admission_number,
                'current_class': str(student.current_class) if student.current_class else None,
                'outstanding': float(outstanding),
            })

        return JsonResponse({
            'success': True,
            'parent_name': parent.full_name,
            'parent_phone': parent.phone_primary,
            'children': children_data,
        })


class StudentImportView(LoginRequiredMixin, OrganizationFilterMixin, RoleRequiredMixin, FormView):
    """View for importing students from Excel file."""
    
    template_name = 'students/student_import.html'
    form_class = StudentImportForm
    success_url = reverse_lazy('students:list')
    allowed_roles = [UserRole.SUPER_ADMIN, UserRole.SCHOOL_ADMIN]
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Import Students from Excel'
        return context
    
    def form_valid(self, form):
        excel_file = form.cleaned_data['excel_file']
        
        # Save uploaded file temporarily
        file_path = default_storage.save(
            f'temp_imports/{excel_file.name}',
            ContentFile(excel_file.read())
        )
        
        try:
            # Get full path
            full_path = default_storage.path(file_path)
            
            # Import students
            stats = StudentService.import_students_from_excel(
                file_path=full_path,
                dry_run=False
            )
            
            # Clean up temp file
            try:
                default_storage.delete(file_path)
            except Exception:
                pass
            
            # Show results
            if stats['errors'] > 0:
                messages.warning(
                    self.request,
                    f"Import completed with {stats['errors']} error(s). "
                    f"Created: {stats['students_created']}, Updated: {stats['students_updated']}, "
                    f"Parents created: {stats['parents_created']}, Skipped: {stats['rows_skipped']}"
                )
                # Show first few errors
                for error in stats['error_details'][:5]:
                    messages.error(self.request, error)
            else:
                messages.success(
                    self.request,
                    f"Successfully imported! Created: {stats['students_created']}, "
                    f"Updated: {stats['students_updated']}, Parents created: {stats['parents_created']}"
                )
            
            return redirect(self.success_url)
            
        except Exception as e:
            # Clean up temp file on error
            try:
                default_storage.delete(file_path)
            except Exception:
                pass
            
            logger.error(f"Error importing students: {str(e)}")
            messages.error(self.request, f'Error importing students: {str(e)}')
            return self.form_invalid(form)


class StudentTemplateDownloadView(LoginRequiredMixin, OrganizationFilterMixin, RoleRequiredMixin, View):
    """Generate and download Excel template for student import."""
    
    allowed_roles = [UserRole.SUPER_ADMIN, UserRole.SCHOOL_ADMIN]
    
    def get(self, request):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment
        from datetime import datetime
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
        
        # Header row
        headers = ['Year', '#', 'Name', 'Class', 'Contacts', 'Total Balance']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add example row
        example_data = ['2025', '2245', 'John Doe Mwangi', 'Grade 1', '+254712345678', '0.00']
        for col, value in enumerate(example_data, start=1):
            ws.cell(row=2, column=col, value=value)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, 3):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Generate response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"student_import_template_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
